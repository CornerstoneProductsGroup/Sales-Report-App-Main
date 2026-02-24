
import pandas as pd
def avg_ignore_zeros_cols(row, cols):
    """
    Average of columns in row ignoring zeros/NaN, and ignoring the earliest week column.
    """
    use_cols = _week_cols_excluding_first(row.to_frame().T, cols)
    vals = []
    for c in use_cols:
        v = row.get(c, np.nan)
        if pd.isna(v):
            continue
        try:
            fv = float(v)
        except Exception:
            continue
        if fv == 0:
            continue
        vals.append(fv)
    return float(np.mean(vals)) if vals else 0.0


def _week_cols_excluding_first(df, week_cols):
    """
    Remove the earliest week column from week_cols (to ignore partial first week).
    Uses parsed week start date from the column name when possible.
    """
    if not week_cols:
        return week_cols
    parsed = [pd.to_datetime(c, errors="coerce") for c in week_cols]
    if all(pd.isna(p) for p in parsed):
        return week_cols[1:] if len(week_cols) > 1 else week_cols
    pairs = [(c, p) for c, p in zip(week_cols, parsed) if pd.notna(p)]
    if not pairs:
        return week_cols[1:] if len(week_cols) > 1 else week_cols
    earliest = min(pairs, key=lambda x: x[1])[0]
    return [c for c in week_cols if c != earliest]


import re
from pathlib import Path
from datetime import date, timedelta

import numpy as np
import pandas as pd
import streamlit as st

MONTH_NAME_TO_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12
}

AVG_WINDOW_OPTIONS = ["4 weeks","5 weeks","6 weeks","7 weeks","8 weeks","9 weeks","10 weeks","11 weeks","12 weeks",
                      "January","February","March","April","May","June","July","August","September","October","November","December"]

def resolve_avg_use(avg_window, use_cols, current_year):
    """Return which week columns to use for averaging.

    Supports:
      - Rolling windows like '8 weeks', '13 weeks', etc.
      - Month names like 'January' (within current_year only)
      - Month+Year like 'January 2026' (explicit year)
    """
    if not use_cols:
        return []

    # Month+Year like 'January 2026'
    if isinstance(avg_window, str):
        mm = re.match(r"^(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})$",
                      avg_window.strip())
        if mm:
            mon, yy = mm.group(1), int(mm.group(2))
            mnum = MONTH_NAME_TO_NUM.get(mon)
            dates = pd.to_datetime(pd.Series(list(use_cols)), errors="coerce")
            mask = (dates.dt.year == int(yy)) & (dates.dt.month == int(mnum))
            return [c for c, ok in zip(use_cols, mask.fillna(False).tolist()) if ok]

    # Month-only choices are within current_year only
    if isinstance(avg_window, str) and avg_window in MONTH_NAME_TO_NUM:
        mnum = MONTH_NAME_TO_NUM[avg_window]
        dates = pd.to_datetime(pd.Series(list(use_cols)), errors="coerce")
        mask = (dates.dt.year == int(current_year)) & (dates.dt.month == int(mnum))
        return [c for c, ok in zip(use_cols, mask.fillna(False).tolist()) if ok]

    # Rolling weeks like '8 weeks'
    if isinstance(avg_window, str) and "week" in avg_window:
        try:
            n = int(avg_window.split()[0])
        except Exception:
            n = 4
        return use_cols[-n:] if len(use_cols) >= n else use_cols

    return use_cols


def load_year_locks() -> set[int]:
    try:
        if DEFAULT_YEAR_LOCKS.exists():
            obj = json.loads(DEFAULT_YEAR_LOCKS.read_text(encoding="utf-8"))
            years = obj.get("locked_years", [])
            return set(int(y) for y in years)
    except Exception:
        pass
    return set()

def save_year_locks(locked_years: set[int]) -> None:
    try:
        DEFAULT_YEAR_LOCKS.write_text(json.dumps({"locked_years": sorted(list(locked_years))}, indent=2), encoding="utf-8")
    except Exception:
        return

def overwrite_sales_rows(target_year: int, retailers: set[str]) -> None:
    """Remove rows from sales_store.csv for the given year + retailers."""
    if not DEFAULT_SALES_STORE.exists():
        return
    try:
        cur = pd.read_csv(DEFAULT_SALES_STORE)
        cur["StartDate"] = pd.to_datetime(cur.get("StartDate"), errors="coerce")
        cur["Retailer"] = cur.get("Retailer", "").map(_normalize_retailer)
        retailers_n = {_normalize_retailer(r) for r in retailers}
        keep = ~((cur["StartDate"].dt.year == int(target_year)) & (cur["Retailer"].isin(retailers_n)))
        cur2 = cur[keep].copy()
        cur2.to_csv(DEFAULT_SALES_STORE, index=False)
    except Exception:
        return

# -------------------------
# Normalization
# -------------------------
def _normalize_retailer(x: str) -> str:
    if x is None:
        return ""
    x = str(x).strip()
    aliases = {
        "home depot": "Depot",
        "depot": "Depot",
        "the home depot": "Depot",
        "lowes": "Lowe's",
        "lowe's": "Lowe's",
        "tractor supply": "Tractor Supply",
        "tsc": "Tractor Supply",
        "amazon": "Amazon",
    }
    key = re.sub(r"\s+", " ", x.lower()).strip()
    return aliases.get(key, x)

def _normalize_sku(x: str) -> str:
    if pd.isna(x):
        return ""
    return str(x).strip()

# -------------------------
# Formatting
# -------------------------
def fmt_currency(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    s = f"${abs(v):,.2f}"
    return f"({s})" if v < 0 else s

def fmt_int(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    return f"{int(round(v)):,.0f}"





def fmt_currency_signed(x) -> str:
    try:
        x = float(x)
    except Exception:
        return str(x)
    return f"-${abs(x):,.0f}" if x < 0 else f"${x:,.0f}"

def fmt_int_signed(x) -> str:
    try:
        x = float(x)
    except Exception:
        return str(x)
    x = int(round(x))
    return f"-{abs(x):,}" if x < 0 else f"{x:,}"

def make_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure dataframe has unique column names (pyarrow requirement)."""
    cols = list(df.columns)
    seen = {}
    new_cols = []
    for c in cols:
        if c not in seen:
            seen[c] = 0
            new_cols.append(c)
        else:
            seen[c] += 1
            new_cols.append(f"{c}.{seen[c]}")
    df2 = df.copy()
    df2.columns = new_cols
    return df2

def fmt_2(x) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return ""
    try:
        v = float(x)
    except Exception:
        return ""
    return f"{v:,.2f}"

def _color(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "inherit"
    try:
        v = float(v)
    except Exception:
        return "inherit"
    if v > 0:
        return "green"
    if v < 0:
        return "red"
    return "inherit"

def _table_height(df: pd.DataFrame, row_px: int = 32, header_px: int = 38, max_px: int = 1100) -> int:
    if df is None:
        return 220
    n = int(df.shape[0])
    h = header_px + (n + 1) * row_px
    return int(min(max(h, 220), max_px))

def style_currency_cols(df: pd.DataFrame, diff_cols=None):
    diff_cols = diff_cols or []
    sty = df.style
    # format all non-first columns as currency
    first = df.columns[0]
    fmt = {c: (lambda v: fmt_currency(v)) for c in df.columns if c != first}
    sty = sty.format(fmt)
    for c in diff_cols:
        if c in df.columns:
            sty = sty.applymap(lambda v: f"color: {_color(v)};", subset=[c])
    return sty

# -------------------------
# Vendor map
# -------------------------
def load_vendor_map(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path)
    cols = {c.lower().strip(): c for c in df.columns}

    def pick(name, fallbacks):
        for k in [name] + fallbacks:
            if k in cols:
                return cols[k]
        return None

    c_retail = pick("retailer", [])
    c_sku = pick("sku", ["item", "item sku"])
    c_vendor = pick("vendor", ["supplier"])
    c_price = pick("price", ["unit price", "cost"])

    out = pd.DataFrame({
        "Retailer": df[c_retail] if c_retail else "",
        "SKU": df[c_sku] if c_sku else "",
        "Vendor": df[c_vendor] if c_vendor else "",
        "Price": df[c_price] if c_price else np.nan,
    })

    out["Retailer"] = out["Retailer"].map(_normalize_retailer)
    out["SKU"] = out["SKU"].map(_normalize_sku)
    out["Vendor"] = out["Vendor"].astype(str).str.strip()
    out["Price"] = pd.to_numeric(out["Price"], errors="coerce")

    # preserve order per retailer
    out["MapOrder"] = 0
    for r, grp in out.groupby("Retailer", sort=False):
        for j, ix in enumerate(grp.index.tolist()):
            out.loc[ix, "MapOrder"] = j

    return out

# -------------------------
# Sales store
# -------------------------
def load_sales_store() -> pd.DataFrame:
    if DEFAULT_SALES_STORE.exists():
        df = pd.read_csv(DEFAULT_SALES_STORE)
        for c in ["StartDate", "EndDate"]:
            if c in df.columns:
                df[c] = pd.to_datetime(df[c], errors="coerce")
        df["Retailer"] = df["Retailer"].map(_normalize_retailer)
        df["SKU"] = df["SKU"].map(_normalize_sku)
        df["Units"] = pd.to_numeric(df["Units"], errors="coerce").fillna(0.0)
        if "UnitPrice" in df.columns:
            df["UnitPrice"] = pd.to_numeric(df["UnitPrice"], errors="coerce")
        else:
            df["UnitPrice"] = np.nan
        return df
    return pd.DataFrame(columns=["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"])



# -------------------------
# Price history (effective dating)
# -------------------------
def _normalize_price_retailer(x):
    x = "" if x is None else str(x).strip()
    if x == "" or x.lower() in {"all","*", "any"}:
        return "*"
    return _normalize_retailer(x)

def load_price_history() -> pd.DataFrame:
    """
    Returns columns: Retailer, SKU, Price, StartDate (datetime64)
    Retailer="*" means applies to all retailers for that SKU.
    """
    if DEFAULT_PRICE_HISTORY.exists():
        ph = pd.read_csv(DEFAULT_PRICE_HISTORY)
        # flexible column names
        colmap = {c.lower(): c for c in ph.columns}
        sku_col = colmap.get("sku") or colmap.get("sku#") or colmap.get("skunumber") or colmap.get("skuid")
        price_col = colmap.get("price") or colmap.get("unitprice") or colmap.get("unit_price")
        date_col = colmap.get("startdate") or colmap.get("start_date") or colmap.get("effective_date") or colmap.get("date")
        ret_col = colmap.get("retailer")

        if sku_col:
            ph["SKU"] = ph[sku_col].map(_normalize_sku)
        else:
            ph["SKU"] = ""
        if price_col:
            ph["Price"] = pd.to_numeric(ph[price_col], errors="coerce")
        else:
            ph["Price"] = np.nan
        if date_col:
            ph["StartDate"] = pd.to_datetime(ph[date_col], errors="coerce")
        else:
            ph["StartDate"] = pd.NaT
        if ret_col:
            ph["Retailer"] = ph[ret_col].map(_normalize_price_retailer)
        else:
            ph["Retailer"] = "*"

        ph = ph[["Retailer","SKU","Price","StartDate"]].dropna(subset=["SKU","Price","StartDate"])
        ph = ph.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
        return ph
    return pd.DataFrame(columns=["Retailer","SKU","Price","StartDate"])

def save_price_history(ph: pd.DataFrame) -> None:
    ph2 = ph.copy()
    ph2["StartDate"] = pd.to_datetime(ph2["StartDate"], errors="coerce")
    ph2 = ph2.dropna(subset=["Retailer","SKU","Price","StartDate"])
    ph2["Retailer"] = ph2["Retailer"].map(_normalize_price_retailer)
    ph2["SKU"] = ph2["SKU"].map(_normalize_sku)
    ph2["Price"] = pd.to_numeric(ph2["Price"], errors="coerce")
    ph2 = ph2.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
    ph2.to_csv(DEFAULT_PRICE_HISTORY, index=False)





# -------------------------
# Caching helpers (performance)
# -------------------------
def _file_mtime(p: Path) -> float:
    try:
        return float(p.stat().st_mtime)
    except Exception:
        return 0.0

@st.cache_data(show_spinner=False)
def cached_vendor_map(path_str: str, mtime: float) -> pd.DataFrame:
    # mtime is included to invalidate cache when the file changes
    return load_vendor_map(Path(path_str))

@st.cache_data(show_spinner=False)
def cached_sales_store(mtime: float) -> pd.DataFrame:
    return load_sales_store()

@st.cache_data(show_spinner=False)
def cached_price_history(mtime: float) -> pd.DataFrame:
    return load_price_history()
def _prepare_price_history_upload(new_rows: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Normalize a price history upload. Returns:
      - normalized rows to consider (SKU, Retailer, StartDate, Price)
      - rows ignored (for reporting)

    Rules:
      - Price blank/NaN => ignored
      - Price <= 0 => ignored (treated as blank)
      - Missing SKU or StartDate => ignored

    Column name matching is forgiving (spaces/underscores/case).
    Accepts: SKU, Retailer, Price, StartDate (or "Start Date", "Effective Date", etc.)
    """
    n = new_rows.copy()

    def norm_key(s: str) -> str:
        s = str(s).strip().lower()
        # keep only alphanumerics to make "Start Date" == "start_date" == "StartDate"
        return re.sub(r"[^a-z0-9]+", "", s)

    cols = {norm_key(c): c for c in n.columns}

    def pick(*keys):
        for k in keys:
            if k in cols:
                return cols[k]
        return None

    sku_col = pick("sku", "sku#", "skunumber", "skuid", "itemsku")
    price_col = pick("price", "unitprice", "unit_price")
    date_col = pick("startdate", "start_date", "startdateeffective", "effectivedate", "effective_date", "start", "date", "startdate1", "startdate2", "startdate3", "startdate4", "startdate5", "startdate6", "startdate7", "startdate8", "startdate9", "startdate10", "startdate11", "startdate12", "startdate13", "startdate14", "startdate15", "startdate16", "startdate17", "startdate18", "startdate19", "startdate20", "startdate21", "startdate22", "startdate23", "startdate24", "startdate25", "startdate26", "startdate27", "startdate28", "startdate29", "startdate30", "startdate31", "startdate32", "startdate33", "startdate34", "startdate35", "startdate36", "startdate37", "startdate38", "startdate39", "startdate40", "startdate41", "startdate42", "startdate43", "startdate44", "startdate45", "startdate46", "startdate47", "startdate48", "startdate49", "startdate50", "startdate51", "startdate52", "startdate53", "startdate54", "startdate55", "startdate56", "startdate57", "startdate58", "startdate59", "startdate60")
    # Common: "start date"
    if date_col is None:
        date_col = pick("startdate", "startdate", "startdate")  # no-op, just for clarity
        date_col = cols.get("startdate") or cols.get("startdate")  # no-op

    # Explicitly support "Start Date" / "Effective Date"
    if date_col is None:
        date_col = pick("startdate", "startdate")  # still none
    if date_col is None:
        date_col = pick("startdate")  # still none

    # Final fallback: try any column that normalizes to "startdate"
    if date_col is None and "startdate" in cols:
        date_col = cols["startdate"]

    ret_col = pick("retailer", "store", "channel")

    if not sku_col or not price_col or not date_col:
        raise ValueError("Price history upload must include columns for SKU, Price, and StartDate (e.g., 'Start Date').")

    norm = pd.DataFrame({
        "SKU": n[sku_col].map(_normalize_sku),
        "Price": pd.to_numeric(n[price_col], errors="coerce"),
        "StartDate": pd.to_datetime(n[date_col], errors="coerce"),
        "Retailer": n[ret_col].map(_normalize_price_retailer) if ret_col else "*",
    })

    ignored = norm.copy()
    ignored["IgnoreReason"] = ""

    mask = ignored["SKU"].isna() | (ignored["SKU"].astype(str).str.strip() == "")
    ignored.loc[mask, "IgnoreReason"] = "Missing SKU"

    mask = ignored["StartDate"].isna()
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Missing StartDate",
        ignored.loc[mask, "IgnoreReason"]
    )

    mask = ignored["Price"].isna()
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Blank Price",
        ignored.loc[mask, "IgnoreReason"]
    )

    mask = (ignored["Price"].notna()) & (ignored["Price"] <= 0)
    ignored.loc[mask, "IgnoreReason"] = np.where(
        ignored.loc[mask, "IgnoreReason"].eq(""),
        "Price <= 0",
        ignored.loc[mask, "IgnoreReason"]
    )

    keep = norm.dropna(subset=["SKU","StartDate","Price"]).copy()
    keep = keep[keep["Price"] > 0].copy()

    ignored = ignored[ignored["IgnoreReason"] != ""].copy()
    keep = keep.reset_index(drop=True)
    ignored = ignored.reset_index(drop=True)
    return keep, ignored


def _price_history_diff(cur: pd.DataFrame, incoming: pd.DataFrame) -> pd.DataFrame:
    """
    Build a diff table for (Retailer, SKU, StartDate).
    Actions: insert/update/noop
    """
    if cur is None or cur.empty:
        base = incoming.copy()
        base["OldPrice"] = np.nan
        base["Action"] = "insert"
        base["PriceDiff"] = np.nan
        return base[["Retailer","SKU","StartDate","OldPrice","Price","PriceDiff","Action"]].sort_values(["Retailer","SKU","StartDate"])

    cur2 = cur.copy()
    cur2["StartDate"] = pd.to_datetime(cur2["StartDate"], errors="coerce")
    inc = incoming.copy()
    inc["StartDate"] = pd.to_datetime(inc["StartDate"], errors="coerce")

    key = ["Retailer","SKU","StartDate"]
    merged = inc.merge(cur2[key + ["Price"]].rename(columns={"Price":"OldPrice"}), on=key, how="left")
    merged["PriceDiff"] = merged["Price"] - merged["OldPrice"]
    merged["Action"] = np.where(merged["OldPrice"].isna(), "insert",
                        np.where(np.isclose(merged["Price"], merged["OldPrice"], equal_nan=True), "noop", "update"))
    return merged[key + ["OldPrice","Price","PriceDiff","Action"]].sort_values(key)

def upsert_price_history(new_rows: pd.DataFrame) -> tuple[int, int, int]:
    """
    Upsert price history with effective dates.
    Returns (inserted, updated, ignored_noop) counts for reporting.
    """
    cur = load_price_history()
    incoming, _ignored = _prepare_price_history_upload(new_rows)

    if incoming.empty:
        return (0, 0, 0)

    diff = _price_history_diff(cur, incoming)
    to_apply = diff[diff["Action"].isin(["insert","update"])].copy()
    noop = int((diff["Action"] == "noop").sum())

    if to_apply.empty:
        return (0, 0, noop)

    apply_rows = to_apply[["Retailer","SKU","StartDate","Price"]].copy()

    merged = pd.concat([cur, apply_rows], ignore_index=True) if (cur is not None and not cur.empty) else apply_rows.copy()
    merged["StartDate"] = pd.to_datetime(merged["StartDate"], errors="coerce")
    merged = merged.dropna(subset=["SKU","Price","StartDate"])
    merged = merged.drop_duplicates(subset=["Retailer","SKU","StartDate"], keep="last")
    merged = merged.sort_values(["Retailer","SKU","StartDate"]).reset_index(drop=True)
    save_price_history(merged)

    inserted = int((diff["Action"] == "insert").sum())
    updated = int((diff["Action"] == "update").sum())
    return (inserted, updated, noop)



def apply_effective_prices(base: pd.DataFrame, vmap: pd.DataFrame, ph: pd.DataFrame) -> pd.DataFrame:
    """
    Hybrid pricing:
      1) If UnitPrice is provided on the weekly sheet, ALWAYS use it (locks history).
      2) Else, use effective-date price history (retailer-specific first, then wildcard '*' retailer for all).
      3) Else, fall back to vendor map Price.

    Notes:
      - merge_asof requires non-null, sorted datetime keys.
    """
    base = base.copy()

    # Ensure expected columns exist
    if "Price" not in base.columns:
        base["Price"] = np.nan
    if "UnitPrice" not in base.columns:
        base["UnitPrice"] = np.nan

    base["StartDate"] = pd.to_datetime(base["StartDate"], errors="coerce")

    # Start with vendor-map price, then let weekly UnitPrice override
    base["PriceEffective"] = base["Price"]
    base["PriceEffective"] = base["UnitPrice"].combine_first(base["PriceEffective"])

    # If no price history, finish
    if ph is None or ph.empty:
        return base

    ph = ph.copy()
    ph["StartDate"] = pd.to_datetime(ph["StartDate"], errors="coerce")
    ph = ph.dropna(subset=["SKU", "StartDate", "Price"]).copy()
    if ph.empty:
        return base

    # Normalize keys
    if "Retailer" not in ph.columns:
        ph["Retailer"] = "*"
    ph["Retailer"] = ph["Retailer"].fillna("*").astype(str).str.strip()
    ph["SKU"] = ph["SKU"].map(_normalize_sku)
    base["SKU"] = base["SKU"].map(_normalize_sku)

    # merge_asof cannot handle NaT in the 'on' key
    base_valid = base[base["StartDate"].notna()].copy()
    base_invalid = base[base["StartDate"].isna()].copy()

    # Retailer-specific history (not '*')
    ph_exact = ph[ph["Retailer"] != "*"].copy()
    ph_star = ph[ph["Retailer"] == "*"].copy()

    # Apply retailer-specific prices
    if not ph_exact.empty and not base_valid.empty:
        b1 = base_valid.sort_values(["StartDate","Retailer","SKU"], kind="mergesort").reset_index(drop=True)
        p1 = ph_exact.sort_values(["StartDate","Retailer","SKU"], kind="mergesort").reset_index(drop=True)

        exact = pd.merge_asof(
            b1,
            p1[["Retailer", "SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_Price"}),
            by=["Retailer", "SKU"],
            on="StartDate",
            direction="backward",
            allow_exact_matches=True,
        )

        # Only use PH_Price when UnitPrice is missing
        exact["PriceEffective"] = exact["UnitPrice"].combine_first(exact["PH_Price"]).combine_first(exact["PriceEffective"])
        exact = exact.drop(columns=["PH_Price"], errors="ignore")
        base_valid = exact

    # Apply wildcard prices to rows still missing PriceEffective (and no UnitPrice)
    if not ph_star.empty and not base_valid.empty:
        missing = base_valid["UnitPrice"].isna() & base_valid["PriceEffective"].isna()
        if missing.any():
            b2 = base_valid.loc[missing].copy()
            b2 = b2.sort_values(["StartDate","SKU"], kind="mergesort").reset_index(drop=True)
            p2 = ph_star.sort_values(["StartDate","SKU"], kind="mergesort").reset_index(drop=True)

            star = pd.merge_asof(
                b2,
                p2[["SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_PriceStar"}),
                by=["SKU"],
                on="StartDate",
                direction="backward",
                allow_exact_matches=True,
            )
            base_valid.loc[missing, "PriceEffective"] = star["PH_PriceStar"].values

    # Final: ensure UnitPrice still wins
    if not base_valid.empty:
        base_valid["PriceEffective"] = base_valid["UnitPrice"].combine_first(base_valid["PriceEffective"])

    # Recombine
    base_out = pd.concat([base_valid, base_invalid], ignore_index=True)
    return base_out

    ph = ph.copy()
    ph["StartDate"] = pd.to_datetime(ph["StartDate"], errors="coerce")
    ph = ph.dropna(subset=["SKU", "StartDate", "Price"]).copy()

    if ph.empty:
        return base

    # Normalize retailer field
    if "Retailer" not in ph.columns:
        ph["Retailer"] = "*"
    ph["Retailer"] = ph["Retailer"].fillna("*").astype(str).str.strip()
    ph["SKU"] = ph["SKU"].map(_normalize_sku)
    base["SKU"] = base["SKU"].map(_normalize_sku)

    # Retailer-specific history (not '*')
    ph_exact = ph[ph["Retailer"] != "*"].copy()
    # Wildcard history applies to all retailers
    ph_star = ph[ph["Retailer"] == "*"].copy()

    # Apply retailer-specific prices using merge_asof
    if not ph_exact.empty:
        b1 = base.sort_values(["Retailer", "SKU", "StartDate"]).reset_index(drop=True)
        p1 = ph_exact.sort_values(["Retailer", "SKU", "StartDate"]).reset_index(drop=True)

        # merge_asof requires both sides sorted by by-keys then on-key
        exact = pd.merge_asof(
            b1,
            p1[["Retailer", "SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_Price"}),
            by=["Retailer", "SKU"],
            on="StartDate",
            direction="backward",
            allow_exact_matches=True,
        )
        base = exact

        # Only use PH_Price when UnitPrice is missing
        base["PriceEffective"] = base["UnitPrice"].combine_first(base["PH_Price"]).combine_first(base["PriceEffective"])
        base = base.drop(columns=["PH_Price"], errors="ignore")

    # Apply wildcard prices for any rows still missing an effective price (and no UnitPrice)
    if not ph_star.empty:
        missing = base["UnitPrice"].isna() & base["PriceEffective"].isna()
        if missing.any():
            b2 = base.loc[missing].copy()
            b2 = b2.sort_values(["SKU", "StartDate"]).reset_index(drop=True)
            p2 = ph_star.sort_values(["SKU", "StartDate"]).reset_index(drop=True)

            star = pd.merge_asof(
                b2,
                p2[["SKU", "StartDate", "Price"]].rename(columns={"Price": "PH_PriceStar"}),
                by=["SKU"],
                on="StartDate",
                direction="backward",
                allow_exact_matches=True,
            )
            base.loc[missing, "PriceEffective"] = star["PH_PriceStar"].values

    # Final: ensure UnitPrice still wins
    base["PriceEffective"] = base["UnitPrice"].combine_first(base["PriceEffective"])

    return base

def upsert_sales(existing: pd.DataFrame, new_rows: pd.DataFrame) -> pd.DataFrame:
    if existing is None or existing.empty:
        return new_rows.copy()
    if new_rows is None or new_rows.empty:
        return existing.copy()

    for c in ["StartDate","EndDate"]:
        if c in existing.columns:
            existing[c] = pd.to_datetime(existing[c], errors="coerce")
        if c in new_rows.columns:
            new_rows[c] = pd.to_datetime(new_rows[c], errors="coerce")

    key_cols = ["Retailer","SKU","StartDate","EndDate","SourceFile"]
    combined = pd.concat([existing, new_rows], ignore_index=True)
    combined = combined.drop_duplicates(subset=key_cols, keep="last")
    return combined

def append_sales_to_store(new_rows: pd.DataFrame) -> None:
    if new_rows is None or new_rows.empty:
        return
    existing = load_sales_store()
    combined = upsert_sales(existing, new_rows)
    combined.to_csv(DEFAULT_SALES_STORE, index=False)

# -------------------------
# Weekly workbook ingestion
# -------------------------
def parse_date_range_from_filename(name: str, year_hint: int):
    n = name.lower()

    m = re.search(r"(\d{4})[-_/](\d{1,2})[-_/](\d{1,2}).*?(?:thru|through|to|–|-).*?(\d{4})[-_/](\d{1,2})[-_/](\d{1,2})", n)
    if m:
        y1, mo1, d1, y2, mo2, d2 = map(int, m.groups())
        return pd.Timestamp(date(y1, mo1, d1)), pd.Timestamp(date(y2, mo2, d2))

    m = re.search(r"(\d{1,2})[-_/](\d{1,2}).*?(?:thru|through|to|–|-).*?(\d{1,2})[-_/](\d{1,2})", n)
    if m:
        mo1, d1, mo2, d2 = map(int, m.groups())
        y = int(year_hint)
        return pd.Timestamp(date(y, mo1, d1)), pd.Timestamp(date(y, mo2, d2))

    return None, None


def read_weekly_workbook(uploaded_file, year: int) -> pd.DataFrame:
    """Read a weekly sales workbook where each sheet is a retailer.
    Expected layout per sheet:
      - Column A: SKU (no header required)
      - Column B: Units
      - Optional Column C: UnitPrice
    NOTE: Some retailers (e.g. Zoro/HomeSelects) may have only a single data row.
    Pandas can sometimes interpret that as header-only depending on the engine,
    so we include an openpyxl fallback to reliably read the first rows.
    """
    import openpyxl
    from io import BytesIO

    # Prefer openpyxl engine for consistency on Streamlit Cloud
    try:
        xls = pd.ExcelFile(uploaded_file, engine="openpyxl")
    except Exception:
        xls = pd.ExcelFile(uploaded_file)

    fname = getattr(uploaded_file, "name", "upload.xlsx")
    sdt, edt = parse_date_range_from_filename(fname, year_hint=year)
    if sdt is None:
        sdt = pd.Timestamp(date.today() - timedelta(days=7))
        edt = pd.Timestamp(date.today())

    # Build an openpyxl workbook once for fallback reads
    wb = None
    try:
        data = uploaded_file.getvalue() if hasattr(uploaded_file, "getvalue") else uploaded_file.read()
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True, read_only=True, keep_links=False)
    except Exception:
        wb = None

    dfs = []
    for sh in xls.sheet_names:
        retailer = _normalize_retailer(sh)

        # Primary read (no headers)
        try:
            raw = pd.read_excel(xls, sheet_name=sh, header=None, engine="openpyxl")
        except Exception:
            raw = pd.read_excel(xls, sheet_name=sh, header=None)

        # Fallback: if pandas returns empty but the sheet has a single row (common for new retailers)
        if (raw is None) or (raw.shape[0] == 0) or (raw.shape[1] < 2):
            if wb is not None and sh in wb.sheetnames:
                ws = wb[sh]
                vals = []
                # read first 500 rows, up to 3 cols, stopping after a run of blanks
                blank_run = 0
                for r in range(1, 501):
                    sku = ws.cell(row=r, column=1).value
                    units = ws.cell(row=r, column=2).value
                    price = ws.cell(row=r, column=3).value
                    if (sku is None or str(sku).strip() == "") and (units is None or str(units).strip() == "") and (price is None or str(price).strip() == ""):
                        blank_run += 1
                        if blank_run >= 20:
                            break
                        continue
                    blank_run = 0
                    vals.append([sku, units, price])
                if vals:
                    raw = pd.DataFrame(vals)
                else:
                    continue
            else:
                continue

        # Keep only first 3 columns
        raw = raw.iloc[:, :3].copy() if raw.shape[1] >= 3 else raw.iloc[:, :2].copy()
        raw.columns = ["SKU", "Units", "UnitPrice"] if raw.shape[1] == 3 else ["SKU", "Units"]

        raw["SKU"] = raw["SKU"].map(_normalize_sku)
        raw["Units"] = pd.to_numeric(raw["Units"], errors="coerce").fillna(0.0)

        if "UnitPrice" in raw.columns:
            raw["UnitPrice"] = pd.to_numeric(raw["UnitPrice"], errors="coerce")
        else:
            raw["UnitPrice"] = np.nan

        raw = raw[raw["SKU"].astype(str).str.strip().ne("")]

        raw["Retailer"] = retailer
        raw["StartDate"] = pd.to_datetime(sdt)
        raw["EndDate"] = pd.to_datetime(edt)
        raw["SourceFile"] = fname
        if "UnitPrice" not in raw.columns:
            raw["UnitPrice"] = np.nan
        dfs.append(raw[["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"]])

    out = pd.concat(dfs, ignore_index=True) if dfs else pd.DataFrame(columns=["Retailer","SKU","Units","UnitPrice","StartDate","EndDate","SourceFile"])
    if not out.empty:
        out["Retailer"] = out["Retailer"].map(_normalize_retailer)
        out["SKU"] = out["SKU"].map(_normalize_sku)
        out["StartDate"] = pd.to_datetime(out["StartDate"], errors="coerce")
        out["EndDate"] = pd.to_datetime(out["EndDate"], errors="coerce")
    return out


# -------------------------
# Year-Overview (YOW) workbook ingestion
# -------------------------
def parse_week_range_header(val, year: int):
    """Parse headers like '1-1 / 1-3' into (StartDate, EndDate) timestamps.
    Accepts a few common variants and handles year-crossing weeks (e.g. '12-29 / 1-2').
    """
    if val is None:
        return (None, None)
    s = str(val).strip()
    if s == "":
        return (None, None)

    # Common: 'M-D / M-D' or 'M/D - M/D'
    m = re.search(r"(\d{1,2})\s*[-/]\s*(\d{1,2})\s*(?:/|to|–|-)+\s*(\d{1,2})\s*[-/]\s*(\d{1,2})", s)
    if not m:
        # Variant with explicit months: '1-1 / 1-3' (same as above but stricter)
        m = re.search(r"(\d{1,2})-(\d{1,2})\s*/\s*(\d{1,2})-(\d{1,2})", s)
    if not m:
        return (None, None)

    mo1, d1, mo2, d2 = map(int, m.groups())
    y1 = int(year)
    y2 = int(year)
    # If the end month is earlier than start month, assume it crosses into next year
    if mo2 < mo1:
        y2 = y1 + 1

    try:
        sdt = pd.Timestamp(y1, mo1, d1)
        edt = pd.Timestamp(y2, mo2, d2)
        return (sdt, edt)
    except Exception:
        return (None, None)

def read_yow_workbook(uploaded_file, year: int) -> pd.DataFrame:
    """Read a Year Overview workbook:
    - One sheet per retailer OR a single sheet where A1 is the retailer name.
    - Row 1 contains week ranges across the top (starting in column B).
    - Column A contains SKUs (starting row 2).
    - Cells contain Units.
    """
    import openpyxl

    fname = getattr(uploaded_file, "name", "yow.xlsx")

    # openpyxl is fastest/most tolerant for wide sheets
    wb = openpyxl.load_workbook(uploaded_file, data_only=True, read_only=True, keep_links=False)

    rows_out = []

    for sh in wb.sheetnames:
        ws = wb[sh]

        # Retailer name: A1 (preferred). If blank, fall back to sheet name.
        retailer_name = ws["A1"].value
        retailer = _normalize_retailer(retailer_name if retailer_name not in [None, ""] else sh)

        # Header row: week ranges from B1 onward until blank
        week_cols = []
        col = 2  # B
        while True:
            v = ws.cell(row=1, column=col).value
            if v is None or str(v).strip() == "":
                break
            sdt, edt = parse_week_range_header(v, year=year)
            if sdt is None:
                # Try interpreting as a date (week start) if someone uses real date headers
                dt = pd.to_datetime(v, errors="coerce")
                if pd.notna(dt):
                    sdt = pd.Timestamp(dt).normalize()
                    edt = sdt + pd.Timedelta(days=6)
                else:
                    # stop if header isn't parseable
                    break
            if edt is None:
                edt = sdt + pd.Timedelta(days=6)
            week_cols.append((col, pd.Timestamp(sdt), pd.Timestamp(edt), str(v).strip()))
            col += 1

        if not week_cols:
            continue

        # Data rows: SKUs down column A from row 2 until blank
        row = 2
        while True:
            sku = ws.cell(row=row, column=1).value
            if sku is None or str(sku).strip() == "":
                break
            sku = _normalize_sku(sku)

            for (cidx, sdt, edt, hdr) in week_cols:
                units = ws.cell(row=row, column=cidx).value
                if units is None or (isinstance(units, str) and units.strip() == ""):
                    continue
                try:
                    u = float(units)
                except Exception:
                    continue
                if np.isnan(u) or u == 0:
                    continue

                rows_out.append({
                    "Retailer": retailer,
                    "SKU": sku,
                    "Units": float(u),
                    "UnitPrice": np.nan,          # use current pricing (vendor map / price history)
                    "StartDate": pd.to_datetime(sdt),
                    "EndDate": pd.to_datetime(edt),
                    "SourceFile": f"{fname}::{sh}",
                })

            row += 1

    out = pd.DataFrame(rows_out)
    if not out.empty:
        out["Retailer"] = out["Retailer"].map(_normalize_retailer)
        out["SKU"] = out["SKU"].map(_normalize_sku)
        out["StartDate"] = pd.to_datetime(out["StartDate"], errors="coerce")
        out["EndDate"] = pd.to_datetime(out["EndDate"], errors="coerce")
        out["Units"] = pd.to_numeric(out["Units"], errors="coerce").fillna(0.0)
        out["UnitPrice"] = pd.to_numeric(out["UnitPrice"], errors="coerce")
    return out

# -------------------------
# Enrichment / metrics
# -------------------------
def enrich_sales(sales: pd.DataFrame, vmap: pd.DataFrame, price_hist: pd.DataFrame | None = None) -> pd.DataFrame:
    s = sales.copy()
    s["Retailer"] = s["Retailer"].map(_normalize_retailer)
    s["SKU"] = s["SKU"].map(_normalize_sku)
    s["Units"] = pd.to_numeric(s["Units"], errors="coerce").fillna(0.0).astype(float)
    s["StartDate"] = pd.to_datetime(s["StartDate"], errors="coerce")
    s["EndDate"] = pd.to_datetime(s["EndDate"], errors="coerce")

    m = vmap[["Retailer","SKU","Vendor","Price","MapOrder"]].copy()
    # Normalize keys; treat blank Retailer in vendor map as wildcard ("*")
    m["Retailer"] = m["Retailer"].fillna("*").map(_normalize_retailer)
    m["Retailer"] = m["Retailer"].replace({"": "*"})
    m["SKU"] = m["SKU"].map(_normalize_sku)
    m["Price"] = pd.to_numeric(m["Price"], errors="coerce")

    # 1) Exact match on Retailer + SKU
    out = s.merge(m[m["Retailer"] != "*"], on=["Retailer","SKU"], how="left")

    # 2) Wildcard match on SKU only for rows still missing pricing/vendor
    wild = m[m["Retailer"] == "*"][["SKU","Vendor","Price","MapOrder"]].drop_duplicates()
    if not wild.empty:
        miss = out["Price"].isna()
        if miss.any():
            out2 = out.loc[miss].merge(wild, on=["SKU"], how="left", suffixes=("", "_w"))
            if "Vendor_w" in out2.columns:
                out.loc[miss, "Vendor"] = out2["Vendor"].combine_first(out2["Vendor_w"]).values
            if "Price_w" in out2.columns:
                out.loc[miss, "Price"] = out2["Price"].combine_first(out2["Price_w"]).values
            if "MapOrder_w" in out2.columns:
                out.loc[miss, "MapOrder"] = out2["MapOrder"].combine_first(out2["MapOrder_w"]).values

    # Apply effective-dated pricing (if provided), otherwise fallback to vendor map price
    ph = price_hist if price_hist is not None else load_price_history()
    out = apply_effective_prices(out, vmap, ph)


    # Compute Sales from Units and effective price (Units-only weekly uploads)
    out["Units"] = pd.to_numeric(out.get("Units", 0), errors="coerce").fillna(0.0)
    out["PriceEffective"] = pd.to_numeric(out.get("PriceEffective", np.nan), errors="coerce")
    out["Sales"] = (out["Units"] * out["PriceEffective"]).fillna(0.0)
    return out



@st.cache_data(show_spinner=False)
def cached_enrich_sales(store_mtime: float, vmap_path: str, vmap_mtime: float, ph_mtime: float) -> pd.DataFrame:
    """
    Cache the expensive enrichment step (merges + effective-dated pricing).
    Cache is invalidated automatically when:
      - sales_store.csv changes (store_mtime)
      - vendor_map.xlsx changes (vmap_mtime)
      - price_history.csv changes (ph_mtime)
    """
    sales_store = load_sales_store()
    vmap = load_vendor_map(Path(vmap_path))
    price_hist = load_price_history()
    return enrich_sales(sales_store, vmap, price_hist)
def wow_mom_metrics(df: pd.DataFrame) -> dict:
    out = {"total_units":0.0,"total_sales":0.0,"wow_units":None,"wow_sales":None,"mom_units":None,"mom_sales":None}
    if df is None or df.empty:
        return out
    d = df.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    out["total_units"] = float(d["Units"].sum())
    out["total_sales"] = float(d["Sales"].fillna(0).sum())

    periods_all = sorted(pd.to_datetime(df_all["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
    periods = sorted(d["StartDate"].dropna().dt.date.unique().tolist())
    if len(periods) >= 1:
        cur_p = periods[-1]
        cur = d[d["StartDate"].dt.date == cur_p]
        cur_u = cur["Units"].sum()
        cur_s = cur["Sales"].fillna(0).sum()
        if len(periods) >= 2:
            prev_p = periods[-2]
            prev = d[d["StartDate"].dt.date == prev_p]
            prev_u = prev["Units"].sum()
            prev_s = prev["Sales"].fillna(0).sum()
        else:
            prev_u = 0.0
            prev_s = 0.0
        out["wow_units"] = float(cur_u - prev_u)
        out["wow_sales"] = float(cur_s - prev_s)

    d["MonthP"] = d["StartDate"].dt.to_period("M")
    months = sorted(d["MonthP"].dropna().unique().tolist())
    if len(months) >= 1:
        cur_m = months[-1]
        cur = d[d["MonthP"] == cur_m]
        cur_u = cur["Units"].sum()
        cur_s = cur["Sales"].fillna(0).sum()
        if len(months) >= 2:
            prev_m = months[-2]
            prev = d[d["MonthP"] == prev_m]
            prev_u = prev["Units"].sum()
            prev_s = prev["Sales"].fillna(0).sum()
        else:
            prev_u = 0.0
            prev_s = 0.0
        out["mom_units"] = float(cur_u - prev_u)
        out["mom_sales"] = float(cur_s - prev_s)

    return out

def month_label(p: pd.Period) -> str:
    return p.to_timestamp().strftime("%B %Y")

APP_TITLE = "Sales Report App"

# -------------------------
# App UI
# -------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")
st.title(APP_TITLE)

with st.sidebar:
    st.header("Data Inputs")
    edit_mode = st.checkbox("Enable Edit Mode (edit Vendor/Price)", value=False)

    this_year = date.today().year
    year = st.selectbox("Year (for filename date parsing)", options=list(range(this_year-3, this_year+2)), index=3)
    view_year = st.selectbox("View Year (dashboard)", options=list(range(this_year-3, this_year+2)), index=3, key="view_year")

    st.subheader("Vendor Map")
    vm_upload = st.file_uploader("Upload Vendor Map (.xlsx)", type=["xlsx"], key="vm_up")
    a, b = st.columns(2)
    with a:
        if st.button("Use uploaded as default", disabled=vm_upload is None):
            DEFAULT_VENDOR_MAP.write_bytes(vm_upload.getbuffer())
            st.success("Saved as default vendor map.")
            st.rerun()
    with b:
        if st.button("Reload"):
            st.rerun()

    st.subheader("Weekly Sales Workbooks")
    wk_uploads = st.file_uploader("Upload weekly sales workbook(s) (.xlsx)", type=["xlsx"], accept_multiple_files=True, key="wk_up")
    if st.button("Ingest uploads", disabled=not wk_uploads):
        all_new = []
        prog = st.progress(0, text="Reading workbooks…")
        total = len(wk_uploads)
        for i, f in enumerate(wk_uploads, start=1):
            try:
                new_rows = read_weekly_workbook(f, year=year)
                if new_rows is not None and not new_rows.empty:
                    all_new.append(new_rows)
            except Exception as e:
                st.error(f"Failed to read {getattr(f, 'name', 'upload')}: {e}")
            prog.progress(i / max(total, 1), text=f"Reading workbooks… ({i}/{total})")
        prog.empty()

        if all_new:
            combined_new = pd.concat(all_new, ignore_index=True)
            append_sales_to_store(combined_new)
            st.success(f"Ingested {len(all_new)} workbook(s) into the sales store.")
        else:
            st.info("No rows found in the uploaded workbooks.")
        st.rerun()

    st.divider()
    if st.button("Clear ALL stored sales data"):
        if DEFAULT_SALES_STORE.exists():
            DEFAULT_SALES_STORE.unlink()
        st.warning("Sales store cleared.")
        st.rerun()


# Ensure view_year exists for downstream tabs
view_year = st.session_state.get('view_year', year)



# Load vendor map (persistent)
BUNDLED_VENDOR_MAP = Path(__file__).parent / "vendor_map.xlsx"

# If a default vendor map hasn't been set yet, seed it from the bundled file in the repo.
try:
    if (not DEFAULT_VENDOR_MAP.exists()) and BUNDLED_VENDOR_MAP.exists():
        DEFAULT_VENDOR_MAP.write_bytes(BUNDLED_VENDOR_MAP.read_bytes())
except Exception:
    pass

if vm_upload is not None:
    tmp = DATA_DIR / "_session_vendor_map.xlsx"
    tmp.write_bytes(vm_upload.getbuffer())
    vmap_path_used = str(tmp)
    vmap = cached_vendor_map(vmap_path_used, _file_mtime(tmp))
elif DEFAULT_VENDOR_MAP.exists():
    vmap_path_used = str(DEFAULT_VENDOR_MAP)
    vmap = cached_vendor_map(vmap_path_used, _file_mtime(DEFAULT_VENDOR_MAP))
else:
    st.info("Upload a vendor map to begin.")
    st.stop()

sales_store = cached_sales_store(_file_mtime(DEFAULT_SALES_STORE))
price_hist = cached_price_history(_file_mtime(DEFAULT_PRICE_HISTORY))
df_all = cached_enrich_sales(_file_mtime(DEFAULT_SALES_STORE), vmap_path_used, _file_mtime(Path(vmap_path_used)), _file_mtime(DEFAULT_PRICE_HISTORY))

# KPIs across top (always current calendar year)
df_kpi = df_all.copy()
df_kpi["StartDate"] = pd.to_datetime(df_kpi["StartDate"], errors="coerce")
df_kpi = df_kpi[df_kpi["StartDate"].dt.year == int(this_year)].copy()

# Apply view-year filter for all reporting tabs
df = df_all.copy()
df["StartDate"] = pd.to_datetime(df["StartDate"], errors="coerce")
df = df[df["StartDate"].dt.year == int(view_year)].copy()

# KPIs across top
m_all = wow_mom_metrics(df_kpi)

st.markdown("## 📊 Overview (All Retailers)")
r1 = st.columns(3)
r2 = st.columns(3)
with r1[0]:
    st.metric("Total Units (YTD)", fmt_int(m_all["total_units"]))
with r1[1]:
    st.metric("Total Sales (YTD)", fmt_currency(m_all["total_sales"]))
with r1[2]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>MoM Units</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['mom_units'])};'>{fmt_int(m_all['mom_units']) if m_all['mom_units'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )
with r2[0]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>MoM Sales</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['mom_sales'])};'>{fmt_currency(m_all['mom_sales']) if m_all['mom_sales'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )
with r2[1]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>WoW Units</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['wow_units'])};'>{fmt_int(m_all['wow_units']) if m_all['wow_units'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )
with r2[2]:
    st.markdown(
        f"<div style='font-size:14px; color: gray;'>WoW Sales</div>"
        f"<div style='font-size:28px; font-weight:600; color:{_color(m_all['wow_sales'])};'>{fmt_currency(m_all['wow_sales']) if m_all['wow_sales'] is not None else '—'}</div>",
        unsafe_allow_html=True
    )

st.divider()


# -------------------------
# Reporting helpers
# -------------------------
def week_labels(df_in: pd.DataFrame) -> list[str]:
    if df_in is None or df_in.empty:
        return []
    w = sorted(pd.to_datetime(df_in["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
    return [pd.Timestamp(x).strftime("%m-%d") for x in w]

def add_week_col(d: pd.DataFrame) -> pd.DataFrame:
    d = d.copy()
    d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
    d["Week"] = d["StartDate"].dt.date
    return d

def nonzero_mean_rowwise(frame: pd.DataFrame) -> pd.Series:
    """Mean across columns, ignoring zeros (treat zeros as missing)."""
    return frame.replace(0, np.nan).mean(axis=1)

def last_n_weeks(df_in: pd.DataFrame, n: int):
    if df_in is None or df_in.empty:
        return []
    w = sorted(pd.to_datetime(df_in["StartDate"], errors="coerce").dropna().dt.date.unique().tolist())
    return w[-n:] if len(w) >= n else w

def safe_div(a, b):
    try:
        if b == 0 or pd.isna(b):
            return np.nan
        return a / b
    except Exception:
        return np.nan

def to_pdf_bytes(title: str, sections: list[tuple[str, list[str]]]) -> bytes:
    """
    Build a simple PDF summary.
    sections: list of (heading, lines[])
    """
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
    except Exception:
        return b""

    import io
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    width, height = letter
    x = 0.75 * inch
    y = height - 0.75 * inch

    c.setFont("Helvetica-Bold", 16)
    c.drawString(x, y, title)
    y -= 0.35 * inch

    for heading, lines in sections:
        if y < 1.0 * inch:
            c.showPage()
            y = height - 0.75 * inch
        c.setFont("Helvetica-Bold", 11)
        c.drawString(x, y, heading)
        y -= 0.22 * inch
        c.setFont("Helvetica", 10)
        for ln in lines:
            if y < 1.0 * inch:
                c.showPage()
                y = height - 0.75 * inch
                c.setFont("Helvetica", 10)
            c.drawString(x, y, str(ln)[:120])
            y -= 0.18 * inch
        y -= 0.10 * inch

    c.save()
    return buf.getvalue()


def render_comparison_retailer_vendor():
        st.subheader("Comparison")

        if df.empty:
            st.info("No sales data yet.")
            return

        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()

        # Build month + year options across ALL years in the store
        d["MonthP"] = d["StartDate"].dt.to_period("M")
        months = sorted(d["MonthP"].unique().tolist())
        month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
        label_to_period = dict(zip(month_labels, months))

        d["Year"] = d["StartDate"].dt.year.astype(int)
        years = sorted(d["Year"].dropna().unique().tolist())

        mode = st.radio(
            "Compare type",
            options=["A vs B (Months)", "A vs B (Years)", "Multi-year (high/low highlight)"],
            index=0,
            horizontal=True,
            key="cmp_mode_v2"
        )

        c1, c2, c3 = st.columns([2, 2, 1])
        with c3:
            by = st.selectbox("Compare by", ["Retailer", "Vendor"], key="cmp_by_v2")

        # Optional limiter list
        if by == "Retailer":
            options = sorted(d["Retailer"].dropna().unique().tolist())
        else:
            options = sorted([v for v in d["Vendor"].dropna().unique().tolist() if str(v).strip()])

        sel = st.multiselect(f"Limit to {by}(s) (optional)", options=options, key="cmp_limit_v2")

        # -------------------------
        # Helper: render A vs B table
        # -------------------------
        def _render_a_vs_b(da: pd.DataFrame, db: pd.DataFrame, label_a: str, label_b: str):
            ga = da.groupby(by, as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
            gb = db.groupby(by, as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))

            out = ga.merge(gb, on=by, how="outer").fillna(0.0)
            out["Units_Diff"] = out["Units_A"] - out["Units_B"]
            out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]
            out["Units_%"] = out["Units_Diff"] / out["Units_B"].replace(0, np.nan)
            out["Sales_%"] = out["Sales_Diff"] / out["Sales_B"].replace(0, np.nan)

            total = {
                by: "TOTAL",
                "Units_A": out["Units_A"].sum(),
                "Sales_A": out["Sales_A"].sum(),
                "Units_B": out["Units_B"].sum(),
                "Sales_B": out["Sales_B"].sum(),
            }
            total["Units_Diff"] = total["Units_A"] - total["Units_B"]
            total["Sales_Diff"] = total["Sales_A"] - total["Sales_B"]
            total["Units_%"] = total["Units_Diff"] / total["Units_B"] if total["Units_B"] else np.nan
            total["Sales_%"] = total["Sales_Diff"] / total["Sales_B"] if total["Sales_B"] else np.nan

            out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

            disp = out[[by,"Units_A","Sales_A","Units_B","Sales_B","Units_Diff","Units_%","Sales_Diff","Sales_%"]].copy()
            disp = disp.rename(columns={
                "Units_A": f"Units ({label_a})",
                "Sales_A": f"Sales ({label_a})",
                "Units_B": f"Units ({label_b})",
                "Sales_B": f"Sales ({label_b})",
            })

            sty = disp.style.format({
                f"Units ({label_a})": fmt_int,
                f"Units ({label_b})": fmt_int,
                "Units_Diff": fmt_int,
                "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                f"Sales ({label_a})": fmt_currency,
                f"Sales ({label_b})": fmt_currency,
                "Sales_Diff": fmt_currency,
                "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
            }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])

            st.dataframe(sty, use_container_width=True, hide_index=True)

        
        def _render_sku_movers(da: pd.DataFrame, db: pd.DataFrame, label_a: str, label_b: str):
            """Top movers by SKU, combined across all retailers in the selection."""
            sa = da.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
            sb = db.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
            out = sa.merge(sb, on="SKU", how="outer").fillna(0.0)
            out["Units_Diff"] = out["Units_A"] - out["Units_B"]
            out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]

            top_up = out.sort_values("Sales_Diff", ascending=False).head(10)
            top_dn = out.sort_values("Sales_Diff", ascending=True).head(10)

            c1, c2 = st.columns(2)
            with c1:
                st.markdown(f"**Top 10 Increased SKUs ({label_a} vs {label_b})**")
                st.dataframe(
                    top_up[["SKU","Units_A","Sales_A","Units_B","Sales_B","Sales_Diff","Units_Diff"]].style.format({
                        "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int,
                        "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency
                    }).applymap(lambda v: f"color: {_color(v)};", subset=["Sales_Diff","Units_Diff"]),
                    use_container_width=True,
                    hide_index=True
                )
            with c2:
                st.markdown(f"**Top 10 Decreased SKUs ({label_a} vs {label_b})**")
                st.dataframe(
                    top_dn[["SKU","Units_A","Sales_A","Units_B","Sales_B","Sales_Diff","Units_Diff"]].style.format({
                        "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int,
                        "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency
                    }).applymap(lambda v: f"color: {_color(v)};", subset=["Sales_Diff","Units_Diff"]),
                    use_container_width=True,
                    hide_index=True
                )
# -------------------------
        # Mode 1: A vs B (Months)
        # -------------------------
        if mode == "A vs B (Months)":
            with c1:
                a_pick = st.multiselect(
                    "Selection A (one or more months)",
                    options=month_labels,
                    default=month_labels[-1:] if month_labels else [],
                    key="cmp_a_months_v2"
                )
            with c2:
                b_pick = st.multiselect(
                    "Selection B (one or more months)",
                    options=month_labels,
                    default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                    key="cmp_b_months_v2"
                )

            a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
            b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

            if not a_periods or not b_periods:
                st.info("Pick at least one month in Selection A and Selection B.")
                return

            da = d[d["MonthP"].isin(a_periods)]
            db = d[d["MonthP"].isin(b_periods)]

            if sel:
                da = da[da[by].isin(sel)]
                db = db[db[by].isin(sel)]

            label_a = " + ".join(a_pick) if a_pick else "A"
            label_b = " + ".join(b_pick) if b_pick else "B"
            _render_a_vs_b(da, db, label_a, label_b)

            st.divider()
            _render_sku_movers(da, db, label_a, label_b)
            return

        # -------------------------
        # Mode 2: A vs B (Years)
        # Example: compare (2023+2024) vs (2024+2025)
        # -------------------------
        if mode == "A vs B (Years)":
            with c1:
                years_a = st.multiselect(
                    "Selection A (one or more years)",
                    options=years,
                    default=years[-2:-1] if len(years) >= 2 else years,
                    key="cmp_years_a_v2"
                )
            with c2:
                years_b = st.multiselect(
                    "Selection B (one or more years)",
                    options=years,
                    default=years[-1:] if years else [],
                    key="cmp_years_b_v2"
                )

            if not years_a or not years_b:
                st.info("Pick at least one year in Selection A and Selection B.")
                return

            da = d[d["Year"].isin([int(y) for y in years_a])]
            db = d[d["Year"].isin([int(y) for y in years_b])]

            if sel:
                da = da[da[by].isin(sel)]
                db = db[db[by].isin(sel)]

            label_a = " + ".join([str(y) for y in years_a])
            label_b = " + ".join([str(y) for y in years_b])
            _render_a_vs_b(da, db, label_a, label_b)

            st.divider()
            _render_sku_movers(da, db, label_a, label_b)
            return

        # -------------------------
        # Mode 3: Multi-year highlight table
        # - pick 2..5 years
        # - show Units_YYYY and Sales_YYYY columns
        # - highlight highest and lowest per row (for Sales columns)
        # -------------------------
        with c1:
            years_pick = st.multiselect(
                "Years to view (pick 2 to 5)",
                options=years,
                default=years[-3:] if len(years) >= 3 else years,
                key="cmp_years_pick_multi_v2"
            )
        with c2:
            metric = st.selectbox(
                "Highlight based on",
                options=["Sales", "Units"],
                index=0,
                key="cmp_multi_metric_v2"
            )

        years_pick = [int(y) for y in years_pick]
        if len(years_pick) < 2:
            st.info("Pick at least two years.")
            return
        years_pick = years_pick[:5]

        dd = d[d["Year"].isin(years_pick)].copy()
        if sel:
            dd = dd[dd[by].isin(sel)]

        pieces = []
        for y in years_pick:
            gy = dd[dd["Year"] == int(y)].groupby(by, as_index=False).agg(**{
                f"Units_{y}": ("Units", "sum"),
                f"Sales_{y}": ("Sales", "sum"),
            })
            pieces.append(gy)

        out = pieces[0]
        for p in pieces[1:]:
            out = out.merge(p, on=by, how="outer")

        out = out.fillna(0.0)

        # Totals row
        total = {by: "TOTAL"}
        for c in out.columns:
            if c == by:
                continue
            total[c] = float(out[c].sum()) if pd.api.types.is_numeric_dtype(out[c]) else ""
        out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

        # Column order
        cols = [by]
        for y in years_pick:
            cols += [f"Units_{y}", f"Sales_{y}"]
        disp = out[cols].copy()

        # Highlight: highest and lowest across selected years for chosen metric
        metric_cols = [f"{metric}_{y}" for y in years_pick if f"{metric}_{y}" in disp.columns]

        # --- Extra insights (trend, CAGR, sparkline) ---
        # Use the selected metric across the chosen years
        spark_chars = ["▁","▂","▃","▄","▅","▆","▇","█"]

        def _sparkline(vals):
            vals = [float(v) if v is not None and not pd.isna(v) else np.nan for v in vals]
            if len(vals) == 0 or all(pd.isna(v) for v in vals):
                return ""
            vmin = np.nanmin(vals)
            vmax = np.nanmax(vals)
            # all equal -> flat line
            if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                return "▁" * len(vals)
            out_s = []
            for v in vals:
                if pd.isna(v):
                    out_s.append(" ")
                    continue
                t = (v - vmin) / (vmax - vmin)
                idx = int(round(t * (len(spark_chars) - 1)))
                idx = max(0, min(len(spark_chars) - 1, idx))
                out_s.append(spark_chars[idx])
            return "".join(out_s)

        def _pct_change(a, b):
            try:
                a = float(a); b = float(b)
            except Exception:
                return np.nan
            if a == 0:
                return np.nan
            return (b - a) / a

        def _cagr(a, b, periods):
            try:
                a = float(a); b = float(b)
            except Exception:
                return np.nan
            if a <= 0 or b <= 0 or periods <= 0:
                return np.nan
            return (b / a) ** (1.0 / periods) - 1.0

        # Build per-row series for chosen metric (exclude TOTAL row)
        metric_year_cols = [(y, f"{metric}_{y}") for y in years_pick if f"{metric}_{y}" in disp.columns]
        if metric_year_cols:
            series_vals = disp[[c for _, c in metric_year_cols]].copy()

            # Sparkline
            disp["Spark"] = series_vals.apply(lambda r: _sparkline(r.tolist()), axis=1)

            # Trend (first -> last)
            first_col = metric_year_cols[0][1]
            last_col = metric_year_cols[-1][1]
            pct = series_vals.apply(lambda r: _pct_change(r[first_col], r[last_col]), axis=1)
            disp["Trend"] = np.where(
                pct.isna(),
                "—",
                np.where(pct > 0, "↑", np.where(pct < 0, "↓", "→"))
            )
            disp["Trend %"] = pct

            # CAGR across (n_years - 1) intervals
            periods = max(1, len(metric_year_cols) - 1)
            disp["CAGR %"] = series_vals.apply(lambda r: _cagr(r[first_col], r[last_col], periods), axis=1)

            # Clear insight columns on TOTAL row (if present)
            try:
                is_total = disp[by].astype(str) == "TOTAL"
                for c in ["Spark", "Trend", "Trend %", "CAGR %"]:
                    if c in disp.columns:
                        disp.loc[is_total, c] = ""
            except Exception:
                pass

        # Move insight columns next to the year columns
        insight_cols = [c for c in ["Spark", "Trend", "Trend %", "CAGR %"] if c in disp.columns]
        if insight_cols:
            disp = disp[[by] + [c for c in disp.columns if c != by and c not in insight_cols] + insight_cols]
        def _hl_minmax(row):
            styles = [""] * len(row)
            # Don't highlight TOTAL row
            if str(row.iloc[0]) == "TOTAL":
                return styles
            vals = []
            idxs = []
            for j, c in enumerate(disp.columns):
                if c in metric_cols:
                    try:
                        v = float(row[c])
                    except Exception:
                        v = np.nan
                    vals.append(v)
                    idxs.append(j)
            if not vals:
                return styles
            vmin = np.nanmin(vals)
            vmax = np.nanmax(vals)
            # If all equal, no highlight
            if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                return styles
            for v, j in zip(vals, idxs):
                if np.isclose(v, vmax):
                    styles[j] = "background-color: rgba(0, 200, 0, 0.18); font-weight: 600;"
                elif np.isclose(v, vmin):
                    styles[j] = "background-color: rgba(220, 0, 0, 0.14);"
            return styles

        fmt = {}
        for c in disp.columns:
            if c.startswith("Units_"):
                fmt[c] = fmt_int
            elif c.startswith("Sales_"):
                fmt[c] = fmt_currency
        if "Trend %" in disp.columns:
            fmt["Trend %"] = lambda v: (f"{float(v)*100:.1f}%" if (v is not None and pd.notna(v) and str(v).strip() not in {"", "—"}) else "—")
        if "CAGR %" in disp.columns:
            fmt["CAGR %"] = lambda v: (f"{float(v)*100:.1f}%" if (v is not None and pd.notna(v) and str(v).strip() not in {"", "—"}) else "—")

        sty = disp.style.format(fmt).apply(_hl_minmax, axis=1)

        st.divider()
        # Movers vs earliest vs latest year (SKU totals across all retailers)
        y0 = int(years_pick[0]); y1 = int(years_pick[-1])
        da = d[d["Year"] == y1].copy()
        db = d[d["Year"] == y0].copy()
        label_a = str(y1)
        label_b = str(y0)
        _render_sku_movers(da, db, label_a, label_b)

        st.dataframe(sty, use_container_width=True, hide_index=True)



def render_comparison_sku():
        st.subheader("SKU Comparison")

        if df.empty:
            st.info("No sales data yet.")
            return

        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()

        d["MonthP"] = d["StartDate"].dt.to_period("M")
        months = sorted(d["MonthP"].unique().tolist())
        month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
        label_to_period = dict(zip(month_labels, months))

        d["Year"] = d["StartDate"].dt.year.astype(int)
        years = sorted(d["Year"].dropna().unique().tolist())

        mode = st.radio(
            "Compare type",
            options=["A vs B (Months)", "A vs B (Years)", "Multi-year (high/low highlight)"],
            index=0,
            horizontal=True,
            key="skucmp_mode_v2"
        )

        # Optional filter by Retailer or Vendor (same as before)
        by = st.selectbox("Filter by", ["Retailer", "Vendor"], key="skucmp_filter_by_v2")
        filt_options = sorted([x for x in d[by].dropna().unique().tolist() if str(x).strip()])
        sel = st.multiselect(f"Limit to {by}(s) (optional)", options=filt_options, key="skucmp_limit_v2")

        if sel:
            d = d[d[by].isin(sel)]

        def _sku_a_vs_b(da, db, label_a, label_b):
            ga = da.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
            gb = db.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))
            out = ga.merge(gb, on="SKU", how="outer").fillna(0.0)
            out["Units_Diff"] = out["Units_A"] - out["Units_B"]
            out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]
            out["Units_%"] = out["Units_Diff"] / out["Units_B"].replace(0, np.nan)
            out["Sales_%"] = out["Sales_Diff"] / out["Sales_B"].replace(0, np.nan)

            out = out.sort_values(["Sales_Diff","Units_Diff"], ascending=False, kind="mergesort")

            total = {
                "SKU": "TOTAL",
                "Units_A": float(out["Units_A"].sum()),
                "Sales_A": float(out["Sales_A"].sum()),
                "Units_B": float(out["Units_B"].sum()),
                "Sales_B": float(out["Sales_B"].sum()),
            }
            total["Units_Diff"] = total["Units_A"] - total["Units_B"]
            total["Sales_Diff"] = total["Sales_A"] - total["Sales_B"]
            total["Units_%"] = (total["Units_Diff"] / total["Units_B"]) if total["Units_B"] else np.nan
            total["Sales_%"] = (total["Sales_Diff"] / total["Sales_B"]) if total["Sales_B"] else np.nan
            out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

            disp = out[["SKU","Units_A","Sales_A","Units_B","Sales_B","Units_Diff","Units_%","Sales_Diff","Sales_%"]].copy()
            disp = disp.rename(columns={
                "Units_A": f"Units ({label_a})",
                "Sales_A": f"Sales ({label_a})",
                "Units_B": f"Units ({label_b})",
                "Sales_B": f"Sales ({label_b})",
            })

            sty = disp.style.format({
                f"Units ({label_a})": fmt_int,
                f"Units ({label_b})": fmt_int,
                "Units_Diff": fmt_int,
                "Units_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                f"Sales ({label_a})": fmt_currency,
                f"Sales ({label_b})": fmt_currency,
                "Sales_Diff": fmt_currency,
                "Sales_%": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
            }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff","Sales_Diff"])

            st.dataframe(sty, use_container_width=True, hide_index=True)

        if mode == "A vs B (Months)":
            c1, c2 = st.columns(2)
            with c1:
                a_pick = st.multiselect("Selection A (one or more months)", options=month_labels,
                                        default=month_labels[-1:] if month_labels else [], key="skucmp_a_months_v2")
            with c2:
                b_pick = st.multiselect("Selection B (one or more months)", options=month_labels,
                                        default=month_labels[-2:-1] if len(month_labels) >= 2 else [], key="skucmp_b_months_v2")
            a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
            b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]
            if not a_periods or not b_periods:
                st.info("Pick at least one month in Selection A and Selection B.")
                return
            da = d[d["MonthP"].isin(a_periods)]
            db = d[d["MonthP"].isin(b_periods)]
            _sku_a_vs_b(da, db, " + ".join(a_pick), " + ".join(b_pick))
            return

        if mode == "A vs B (Years)":
            c1, c2 = st.columns(2)
            with c1:
                years_a = st.multiselect("Selection A (one or more years)", options=years,
                                         default=years[-2:-1] if len(years) >= 2 else years, key="skucmp_years_a_v2")
            with c2:
                years_b = st.multiselect("Selection B (one or more years)", options=years,
                                         default=years[-1:] if years else [], key="skucmp_years_b_v2")
            if not years_a or not years_b:
                st.info("Pick at least one year in Selection A and Selection B.")
                return
            da = d[d["Year"].isin([int(y) for y in years_a])]
            db = d[d["Year"].isin([int(y) for y in years_b])]
            _sku_a_vs_b(da, db, " + ".join([str(y) for y in years_a]), " + ".join([str(y) for y in years_b]))
            return

        # Multi-year highlight (SKU totals)
        years_pick = st.multiselect("Years to view (pick 2 to 5)", options=years,
                                    default=years[-3:] if len(years) >= 3 else years, key="skucmp_years_pick_multi_v2")
        years_pick = [int(y) for y in years_pick][:5]
        metric = st.selectbox("Highlight based on", options=["Sales", "Units"], index=0, key="skucmp_multi_metric_v2")
        if len(years_pick) < 2:
            st.info("Pick at least two years.")
            return

        dd = d[d["Year"].isin(years_pick)].copy()
        pieces = []
        for y in years_pick:
            gy = dd[dd["Year"] == int(y)].groupby("SKU", as_index=False).agg(**{
                f"Units_{y}": ("Units", "sum"),
                f"Sales_{y}": ("Sales", "sum"),
            })
            pieces.append(gy)
        out = pieces[0]
        for p in pieces[1:]:
            out = out.merge(p, on="SKU", how="outer")
        out = out.fillna(0.0)

        # Totals row
        total = {"SKU": "TOTAL"}
        for c in out.columns:
            if c == "SKU":
                continue
            total[c] = float(out[c].sum()) if pd.api.types.is_numeric_dtype(out[c]) else ""
        out = pd.concat([out, pd.DataFrame([total])], ignore_index=True)

        cols = ["SKU"]
        for y in years_pick:
            cols += [f"Units_{y}", f"Sales_{y}"]
        disp = out[cols].copy()

        metric_cols = [f"{metric}_{y}" for y in years_pick if f"{metric}_{y}" in disp.columns]

        def _hl_minmax(row):
            styles = [""] * len(row)
            if str(row.iloc[0]) == "TOTAL":
                return styles
            vals = []
            idxs = []
            for j, c in enumerate(disp.columns):
                if c in metric_cols:
                    try:
                        v = float(row[c])
                    except Exception:
                        v = np.nan
                    vals.append(v)
                    idxs.append(j)
            if not vals:
                return styles
            vmin = np.nanmin(vals)
            vmax = np.nanmax(vals)
            if np.isnan(vmin) or np.isnan(vmax) or np.isclose(vmin, vmax):
                return styles
            for v, j in zip(vals, idxs):
                if np.isclose(v, vmax):
                    styles[j] = "background-color: rgba(0, 200, 0, 0.18); font-weight: 600;"
                elif np.isclose(v, vmin):
                    styles[j] = "background-color: rgba(220, 0, 0, 0.14);"
            return styles

        fmt = {}
        for c in disp.columns:
            if c.startswith("Units_"):
                fmt[c] = fmt_int
            elif c.startswith("Sales_"):
                fmt[c] = fmt_currency

        st.dataframe(disp.style.format(fmt).apply(_hl_minmax, axis=1), use_container_width=True, hide_index=True)


def render_sku_health():
        st.subheader("SKU Health Score")

        if df_all.empty:
            st.info("No sales data yet.")
            return

        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()

        d["Year"] = d["StartDate"].dt.year.astype(int)
        d["Month"] = d["StartDate"].dt.month.astype(int)
        d["MonthP"] = d["StartDate"].dt.to_period("M")

        compare_mode = st.selectbox(
            "Compare mode",
            options=["Year vs Year", "Month vs Month (multi-month)"],
            index=0,
            key="sh_compare_mode"
        )

        basis = st.radio("Primary basis", options=["Sales", "Units"], index=0, horizontal=True, key="sh_basis")

        # Shared filters
        f1, f2, f3, f4 = st.columns([2, 2, 1, 1])
        with f1:
            vendor_filter = st.multiselect(
                "Vendor filter (optional)",
                options=sorted([x for x in d["Vendor"].dropna().unique().tolist() if str(x).strip()]),
                key="sh_vendor_filter"
            )
        with f2:
            retailer_filter = st.multiselect(
                "Retailer filter (optional)",
                options=sorted([x for x in d["Retailer"].dropna().unique().tolist() if str(x).strip()]),
                key="sh_retailer_filter"
            )
        with f3:
            top_n = st.number_input("Top N", min_value=20, max_value=2000, value=200, step=20, key="sh_topn")
        with f4:
            status_pick = st.multiselect(
                "Status",
                options=["🔥 Strong","📈 Growing","⚠ Watch","❌ At Risk"],
                default=["🔥 Strong","📈 Growing","⚠ Watch","❌ At Risk"],
                key="sh_status"
            )

        dd = d.copy()
        if vendor_filter:
            dd = dd[dd["Vendor"].isin(vendor_filter)]
        if retailer_filter:
            dd = dd[dd["Retailer"].isin(retailer_filter)]

        # Build A vs B selections
        if compare_mode == "Year vs Year":
            years = sorted(dd["Year"].unique().tolist())
            c1, c2, c3 = st.columns([1, 1, 2])
            with c1:
                base_year = st.selectbox("Base year", options=years, index=max(0, len(years)-2), key="sh_base")
            with c2:
                comp_year = st.selectbox("Compare to", options=years, index=len(years)-1 if years else 0, key="sh_comp")
            with c3:
                pmode = st.selectbox("Period", options=["Full year", "Specific months"], index=0, key="sh_period_mode")

            sel_months = list(range(1,13))
            if pmode == "Specific months":
                month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
                month_list = [month_name[i] for i in range(1,13)]
                sel_names = st.multiselect("Months", options=month_list, default=[month_list[0]], key="sh_months_pick")
                sel_months = [k for k,v in month_name.items() if v in sel_names]

            a = dd[(dd["Year"] == int(base_year)) & (dd["Month"].isin(sel_months))].copy()
            b = dd[(dd["Year"] == int(comp_year)) & (dd["Month"].isin(sel_months))].copy()

            a_label = str(base_year)
            b_label = str(comp_year)

        else:
            # Month vs Month (can be same year or different years)
            months = sorted(dd["MonthP"].unique().tolist())
            month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
            label_to_period = dict(zip(month_labels, months))

            c1, c2 = st.columns(2)
            with c1:
                a_pick = st.multiselect(
                    "Selection A (one or more months)",
                    options=month_labels,
                    default=month_labels[-1:] if month_labels else [],
                    key="sh_mm_a"
                )
            with c2:
                b_pick = st.multiselect(
                    "Selection B (one or more months)",
                    options=month_labels,
                    default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                    key="sh_mm_b"
                )

            a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
            b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

            if (not a_periods) or (not b_periods):
                st.info("Pick at least one month in Selection A and Selection B.")
                return

            a = dd[dd["MonthP"].isin(a_periods)].copy()
            b = dd[dd["MonthP"].isin(b_periods)].copy()

            a_label = "Selection A"
            b_label = "Selection B"

        ga = a.groupby("SKU", as_index=False).agg(Sales_A=("Sales","sum"), Units_A=("Units","sum"))
        gb = b.groupby("SKU", as_index=False).agg(Sales_B=("Sales","sum"), Units_B=("Units","sum"))
        out = ga.merge(gb, on="SKU", how="outer").fillna(0.0)

        # Coverage context (based on B selection)
        cov = b.groupby("SKU", as_index=False).agg(Retailers=("Retailer","nunique"), ActiveWeeks=("StartDate","nunique"))
        out = out.merge(cov, on="SKU", how="left").fillna({"Retailers": 0, "ActiveWeeks": 0})

        out["Δ Sales"] = out["Sales_B"] - out["Sales_A"]
        out["Δ Units"] = out["Units_B"] - out["Units_A"]
        out["Sales %"] = out["Δ Sales"] / out["Sales_A"].replace(0, np.nan)
        out["Units %"] = out["Δ Units"] / out["Units_A"].replace(0, np.nan)

        out["Score"] = out["Δ Sales"] if basis == "Sales" else out["Δ Units"]

        def _status(row):
            a0 = float(row["Sales_A"] if basis=="Sales" else row["Units_A"])
            b0 = float(row["Sales_B"] if basis=="Sales" else row["Units_B"])
            delta = b0 - a0
            if a0 == 0 and b0 > 0:
                return "📈 Growing"
            if a0 > 0 and b0 == 0:
                return "❌ At Risk"
            if delta > 0:
                return "🔥 Strong"
            if delta < 0:
                return "⚠ Watch"
            return "⚠ Watch"

        out["Status"] = out.apply(_status, axis=1)
        out = out[out["Status"].isin(status_pick)].copy()
        out = out.sort_values("Score", ascending=False, kind="mergesort").head(int(top_n))

        # Vendor lookup
        try:
            if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                out = out.merge(vmap[["SKU","Vendor"]].drop_duplicates(), on="SKU", how="left")
        except Exception:
            pass

        cols = ["SKU"] + (["Vendor"] if "Vendor" in out.columns else []) + ["Status","Sales_A","Sales_B","Δ Sales","Sales %","Units_A","Units_B","Δ Units","Units %","Retailers","ActiveWeeks"]
        disp = out[cols].copy()
        disp = disp.rename(columns={
            "Sales_A": a_label,
            "Sales_B": b_label,
            "Units_A": f"Units {a_label}",
            "Units_B": f"Units {b_label}",
        })
        disp = make_unique_columns(disp)

        st.dataframe(
            disp.style.format({
                a_label: fmt_currency,
                b_label: fmt_currency,
                "Δ Sales": fmt_currency,
                "Sales %": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                f"Units {a_label}": fmt_int,
                f"Units {b_label}": fmt_int,
                "Δ Units": fmt_int,
                "Units %": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                "Retailers": fmt_int,
                "ActiveWeeks": fmt_int,
            }),
            use_container_width=True,
            hide_index=True,
            height=_table_height(disp, max_px=1200)
        )

def render_lost_sales():
        st.subheader("Lost Sales Detector")

        if df_all.empty:
            st.info("No sales data yet.")
            return

        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)
        d["Month"] = d["StartDate"].dt.month.astype(int)

        years = sorted(d["Year"].unique().tolist())
        month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
        month_list = [month_name[i] for i in range(1,13)]

        c1, c2, c3 = st.columns([1, 1, 2])
        with c1:
            base_year = st.selectbox("Base year", options=years, index=max(0, len(years)-2), key="ls_base")
        with c2:
            comp_year = st.selectbox("Compare to", options=years, index=len(years)-1, key="ls_comp")
        with c3:
            basis = st.radio("Basis", options=["Sales", "Units"], index=0, horizontal=True, key="ls_basis")

        pmode = st.selectbox("Period", options=["Full year", "Specific months"], index=0, key="ls_period_mode")
        sel_months = list(range(1,13))
        if pmode == "Specific months":
            sel_names = st.multiselect("Months", options=month_list, default=[month_list[0]], key="ls_months_pick")
            sel_months = [k for k,v in month_name.items() if v in sel_names]

        value_col = "Sales" if basis == "Sales" else "Units"

        a = d[(d["Year"] == int(base_year)) & (d["Month"].isin(sel_months))].copy()
        b = d[(d["Year"] == int(comp_year)) & (d["Month"].isin(sel_months))].copy()

        ga = a.groupby("SKU", as_index=False).agg(A=(value_col,"sum"))
        gb = b.groupby("SKU", as_index=False).agg(B=(value_col,"sum"))
        sku = ga.merge(gb, on="SKU", how="outer").fillna(0.0)
        sku["Delta"] = sku["B"] - sku["A"]
        sku["Pct"] = sku["Delta"] / sku["A"].replace(0, np.nan)

        lost = sku[(sku["A"] > 0) & (sku["B"] == 0)].copy().sort_values("A", ascending=False).head(200)
        drops = sku[(sku["A"] > 0) & (sku["B"] > 0) & (sku["Delta"] < 0)].copy().sort_values("Delta").head(200)

        ra = a.groupby(["SKU","Retailer"], as_index=False).agg(A=(value_col,"sum"))
        rb = b.groupby(["SKU","Retailer"], as_index=False).agg(B=(value_col,"sum"))
        rr = ra.merge(rb, on=["SKU","Retailer"], how="outer").fillna(0.0)
        rr["Delta"] = rr["B"] - rr["A"]
        lost_retail = rr[(rr["A"] > 0) & (rr["B"] == 0)].copy().sort_values("A", ascending=False).head(300)

        try:
            if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                vend = vmap[["SKU","Vendor"]].drop_duplicates()
                lost = lost.merge(vend, on="SKU", how="left")
                drops = drops.merge(vend, on="SKU", how="left")
                lost_retail = lost_retail.merge(vend, on="SKU", how="left")
        except Exception:
            pass

        def _fmt(v):
            return fmt_currency(v) if value_col == "Sales" else fmt_int(v)

        st.markdown("### Lost SKUs (sold in base period, zero in compare period)")
        lost_disp = lost[["SKU"] + (["Vendor"] if "Vendor" in lost.columns else []) + ["A"]].copy().rename(columns={"A": str(base_year)})
        lost_disp = make_unique_columns(lost_disp)
        st.dataframe(lost_disp.style.format({str(base_year): _fmt}), use_container_width=True, hide_index=True, height=650)

        # Gained SKUs: 0 in base period, >0 in compare period
        gained = out[(out["Sales_A"] == 0) & (out["Sales_B"] > 0)].copy()
        st.markdown("### Gained SKUs")
        st.dataframe(
            gained.head(500).style.format({
                "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Δ Sales": fmt_currency,
                "Units_A": fmt_int, "Units_B": fmt_int, "Δ Units": fmt_int,
            }).applymap(lambda v: f"color: {_color(v)};", subset=["Δ Sales","Δ Units"]),
            use_container_width=True,
            hide_index=True
        )

        # Net change summary (Gained - Lost)
        lost_tot_units = float(lost["Δ Units"].sum()) if "Δ Units" in lost.columns else 0.0
        lost_tot_sales = float(lost["Δ Sales"].sum()) if "Δ Sales" in lost.columns else 0.0
        gained_tot_units = float(gained["Δ Units"].sum()) if "Δ Units" in gained.columns else 0.0
        gained_tot_sales = float(gained["Δ Sales"].sum()) if "Δ Sales" in gained.columns else 0.0

        net_units = gained_tot_units + lost_tot_units
        net_sales = gained_tot_sales + lost_tot_sales

        st.markdown(
            f"**Net Units (Gained + Lost):** <span style='color:{_color(net_units)}; font-weight:700;'>{fmt_int(net_units)}</span> &nbsp;&nbsp; "
            f"**Net Sales (Gained + Lost):** <span style='color:{_color(net_sales)}; font-weight:700;'>{fmt_currency(net_sales)}</span>",
            unsafe_allow_html=True
        )

        st.markdown("### Biggest declines (still selling, but down)")
        drops_disp = drops[["SKU"] + (["Vendor"] if "Vendor" in drops.columns else []) + ["A","B","Delta","Pct"]].copy()
        drops_disp = drops_disp.rename(columns={"A": str(base_year), "B": str(comp_year)})
        drops_disp = make_unique_columns(drops_disp)
        st.dataframe(
            drops_disp.style.format({
                str(base_year): _fmt,
                str(comp_year): _fmt,
                "Delta": _fmt,
                "Pct": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
            }),
            use_container_width=True,
            hide_index=True,
            height=650
        )

        st.markdown("### Lost retailers for specific SKUs")
        lost_retail_disp = lost_retail[["SKU","Retailer"] + (["Vendor"] if "Vendor" in lost_retail.columns else []) + ["A"]].copy()
        lost_retail_disp = lost_retail_disp.rename(columns={"A": str(base_year)})
        lost_retail_disp = make_unique_columns(lost_retail_disp)
        st.dataframe(lost_retail_disp.style.format({str(base_year): _fmt}), use_container_width=True, hide_index=True, height=700)

def render_data_inventory():
        st.subheader("Data Inventory")

        if df_all.empty:
            st.info("No sales data yet.")
        else:
            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()
            d["Year"] = d["StartDate"].dt.year.astype(int)

            st.markdown("### By year")
            by_year = d.groupby("Year", as_index=False).agg(
                Units=("Units","sum"),
                Sales=("Sales","sum"),
                Retailers=("Retailer","nunique"),
                Vendors=("Vendor","nunique"),
                SKUs=("SKU","nunique"),
            ).sort_values("Year", ascending=False)
            st.dataframe(by_year.style.format({
                "Units": fmt_int, "Sales": fmt_currency,
                "Retailers": fmt_int, "Vendors": fmt_int, "SKUs": fmt_int
            }), use_container_width=True, hide_index=True)

            st.markdown("### By retailer (selected year)")
            years = sorted(d["Year"].unique().tolist())
            sel_y = st.selectbox("Year", options=years, index=len(years)-1, key="inv_year")
            dy = d[d["Year"] == int(sel_y)].copy()
            if "SourceFile" not in dy.columns:
                dy["SourceFile"] = ""
            by_ret = dy.groupby("Retailer", as_index=False).agg(
                Units=("Units","sum"),
                Sales=("Sales","sum"),
                SKUs=("SKU","nunique"),
                Sources=("SourceFile","nunique"),
            ).sort_values("Sales", ascending=False)
            st.dataframe(by_ret.style.format({
                "Units": fmt_int, "Sales": fmt_currency, "SKUs": fmt_int, "Sources": fmt_int
            }), use_container_width=True, height=_table_height(by_ret, max_px=900), hide_index=True)

            st.markdown("### By source file (selected year)")
            by_src = dy.groupby("SourceFile", as_index=False).agg(
                Units=("Units","sum"),
                Sales=("Sales","sum"),
                Retailers=("Retailer","nunique"),
                SKUs=("SKU","nunique"),
            ).sort_values("Sales", ascending=False)
            st.dataframe(by_src.style.format({
                "Units": fmt_int, "Sales": fmt_currency, "Retailers": fmt_int, "SKUs": fmt_int
            }), use_container_width=True, height=_table_height(by_src, max_px=900), hide_index=True)




    # -------------------------
    # Insights & Alerts
    # -------------------------


def render_edit_vendor_map():
        st.subheader("Edit Vendor Map")
        st.caption("Edit Vendor and Price. Click Save to update the default vendor map file used by the app.")
        vmap_disp = vmap[["Retailer","SKU","Vendor","Price","MapOrder"]].copy().sort_values(["Retailer","MapOrder"])
        show = vmap_disp.drop(columns=["MapOrder"]).copy()

        if edit_mode:
            edited = st.data_editor(show, use_container_width=True, hide_index=True, num_rows="dynamic")
            if st.button("Save Vendor Map"):
                updated = edited.copy()
                updated["Retailer"] = updated["Retailer"].map(_normalize_retailer)
                updated["SKU"] = updated["SKU"].map(_normalize_sku)
                updated["Vendor"] = updated["Vendor"].astype(str).str.strip()
                updated["Price"] = pd.to_numeric(updated["Price"], errors="coerce")

                # MapOrder based on current row order per retailer
                updated["MapOrder"] = 0
                for r, grp in updated.groupby("Retailer", sort=False):
                    for j, ix in enumerate(grp.index.tolist()):
                        updated.loc[ix, "MapOrder"] = j

                updated.to_excel(DEFAULT_VENDOR_MAP, index=False)
                st.success("Saved vendor map. Reloading…")
                st.rerun()
        else:
            st.info("Turn on Edit Mode in the sidebar to edit.")
            st.dataframe(show, use_container_width=True, height=_table_height(show, max_px=1400), hide_index=True)

    # Backup / Restore


def render_backup_restore():
        st.subheader("Backup / Restore")

        st.markdown("### Backup files")
        c1, c2, c3 = st.columns(3)

        with c1:
            st.markdown("#### Sales database")
            if DEFAULT_SALES_STORE.exists():
                st.download_button("Download sales_store.csv", data=DEFAULT_SALES_STORE.read_bytes(), file_name="sales_store.csv", mime="text/csv")
            else:
                st.info("No sales_store.csv yet.")

            up = st.file_uploader("Restore sales_store.csv", type=["csv"], key="restore_sales_csv")
            if st.button("Restore sales_store.csv", disabled=up is None, key="btn_restore_sales"):
                DEFAULT_SALES_STORE.write_bytes(up.getbuffer())
                st.success("Restored sales_store.csv. Reloading…")
                st.rerun()

        with c2:
            st.markdown("#### Vendor map")
            if DEFAULT_VENDOR_MAP.exists():
                st.download_button("Download vendor_map.xlsx", data=DEFAULT_VENDOR_MAP.read_bytes(), file_name="vendor_map.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            else:
                st.info("No vendor_map.xlsx yet.")

            up2 = st.file_uploader("Restore vendor_map.xlsx", type=["xlsx"], key="restore_vm_xlsx")
            if st.button("Restore vendor_map.xlsx", disabled=up2 is None, key="btn_restore_vm"):
                DEFAULT_VENDOR_MAP.write_bytes(up2.getbuffer())
                st.success("Restored vendor_map.xlsx. Reloading…")
                st.rerun()

        with c3:
            st.markdown("#### Price history")
            if DEFAULT_PRICE_HISTORY.exists():
                st.download_button("Download price_history.csv", data=DEFAULT_PRICE_HISTORY.read_bytes(), file_name="price_history.csv", mime="text/csv")
            else:
                st.info("No price_history.csv yet.")

            up3 = st.file_uploader("Restore price_history.csv", type=["csv"], key="restore_ph_csv")
            if st.button("Restore price_history.csv", disabled=up3 is None, key="btn_restore_ph"):
                DEFAULT_PRICE_HISTORY.write_bytes(up3.getbuffer())
                st.success("Restored price_history.csv. Reloading…")
                st.rerun()

        st.markdown("#### Year locks")
        if DEFAULT_YEAR_LOCKS.exists():
            st.download_button("Download year_locks.json", data=DEFAULT_YEAR_LOCKS.read_bytes(), file_name="year_locks.json", mime="application/json")
        else:
            st.info("No year locks saved yet.")

        up4 = st.file_uploader("Restore year_locks.json", type=["json"], key="restore_year_locks")
        if st.button("Restore year_locks.json", disabled=up4 is None, key="btn_restore_year_locks"):
            DEFAULT_YEAR_LOCKS.write_bytes(up4.getbuffer())
            st.success("Restored year locks. Reloading…")
            st.rerun()

        st.divider()

        st.markdown("### Price changes (effective date)")
        st.caption("Upload a sheet with SKU + Price + StartDate. Optional Retailer column. Prices apply from StartDate forward and never change earlier weeks.")

        tmpl = pd.DataFrame([
            {"Retailer":"*", "SKU":"ABC123", "Price": 19.99, "StartDate":"2026-02-01"},
            {"Retailer":"home depot", "SKU":"XYZ999", "Price": 24.99, "StartDate":"2026-03-15"},
        ])
        st.download_button("Download template CSV", data=tmpl.to_csv(index=False).encode("utf-8"),
                           file_name="price_history_template.csv", mime="text/csv")

        ph_up = st.file_uploader("Upload price history (CSV or Excel)", type=["csv","xlsx"], key="ph_upload")
        if ph_up is not None:
            try:
                if ph_up.name.lower().endswith(".csv"):
                    ph_new = pd.read_csv(ph_up)
                else:
                    ph_new = pd.read_excel(ph_up)

                st.markdown("#### Preview upload")
                st.dataframe(ph_new.head(50), use_container_width=True, hide_index=True)

                # Normalize + ignore blanks safely
                cur_ph = load_price_history()
                incoming, ignored = _prepare_price_history_upload(ph_new)
                diff = _price_history_diff(cur_ph, incoming)

                st.divider()
                st.markdown("#### What will change")
                c1, c2, c3, c4 = st.columns(4)
                c1.metric("Rows uploaded", int(len(ph_new)))
                c2.metric("Rows ignored (blank/invalid)", int(len(ignored)))
                c3.metric("Inserts", int((diff["Action"] == "insert").sum()) if not diff.empty else 0)
                c4.metric("Updates", int((diff["Action"] == "update").sum()) if not diff.empty else 0)

                show_diff = diff.copy()
                if not show_diff.empty:
                    show_diff["StartDate"] = pd.to_datetime(show_diff["StartDate"], errors="coerce").dt.date
                    sty = show_diff.style.format({
                        "OldPrice": lambda v: fmt_currency(v) if pd.notna(v) else "—",
                        "Price": lambda v: fmt_currency(v),
                        "PriceDiff": lambda v: fmt_currency(v) if pd.notna(v) else "—",
                    }).applymap(lambda v: "font-weight:700;" if str(v) in ["insert","update"] else "", subset=["Action"])
                    st.dataframe(sty, use_container_width=True, height=_table_height(show_diff, max_px=900), hide_index=True)

                    st.download_button("Download change preview (CSV)", data=show_diff.to_csv(index=False).encode("utf-8"),
                                       file_name="price_history_changes_preview.csv", mime="text/csv")
                else:
                    st.info("No valid rows found in this upload (all prices were blank/invalid).")

                if not ignored.empty:
                    st.markdown("#### Ignored rows")
                    ign = ignored.copy()
                    ign["StartDate"] = pd.to_datetime(ign["StartDate"], errors="coerce").dt.date
                    st.dataframe(ign.head(200), use_container_width=True, height=_table_height(ign, max_px=600), hide_index=True)
                    st.download_button("Download ignored rows (CSV)", data=ign.to_csv(index=False).encode("utf-8"),
                                       file_name="price_history_ignored_rows.csv", mime="text/csv")

                if st.button("Apply price changes", key="btn_apply_prices"):
                    ins, upd, noop = upsert_price_history(ph_new)
                    st.success(f"Price history updated. Inserts: {ins}, Updates: {upd}, Unchanged: {noop}. Reloading…")
                    st.rerun()
            except Exception as e:
                st.error(f"Could not read this file: {e}")

        if DEFAULT_PRICE_HISTORY.exists():
            if st.button("Clear ALL price history", key="btn_clear_ph"):
                DEFAULT_PRICE_HISTORY.unlink(missing_ok=True)
                st.success("Cleared. Reloading…")
                st.rerun()

        st.divider()

        st.markdown("### Export enriched sales")
        if not df.empty:
            ex = df.copy()
            ex["StartDate"] = pd.to_datetime(ex["StartDate"], errors="coerce").dt.strftime("%Y-%m-%d")
            ex["EndDate"] = pd.to_datetime(ex["EndDate"], errors="coerce").dt.strftime("%Y-%m-%d")
            st.download_button("Download enriched_sales.csv", data=ex.to_csv(index=False).encode("utf-8"),
                               file_name="enriched_sales.csv", mime="text/csv")
        else:
            st.info("No sales yet.")



    # -------------------------
    # Bulk Data Upload
    # -------------------------


def render_bulk_data_upload():
    st.subheader("Bulk Data Upload (Multi-week / Multi-month)")

    st.markdown(
        """
        Use this when you get a **wide** retailer file (not week-by-week uploads).

        Expected format:
        - One sheet per retailer (or retailer name in cell **A1**)
        - Column **A** = SKU (starting row 2)
        - Row **1** from column **B** onward = week ranges (example: `1-1 / 1-3`)
        - Cells = Units sold for that SKU in that week
        - Sales uses your **current pricing** (Vendor Map / Price History). `UnitPrice` is left blank.
        """
    )

    locked_years = load_year_locks()
    years_opt = list(range(this_year - 6, this_year + 2))

    st.markdown("### Year locks")
    cL1, cL2 = st.columns([2, 1])
    with cL1:
        lock_pick = st.multiselect("Locked years (prevent edits)", options=years_opt, default=sorted(list(locked_years)), key="lock_pick")
    with cL2:
        if st.button("Save locks", key="btn_save_locks"):
            save_year_locks(set(int(y) for y in lock_pick))
            st.success("Saved year locks.")
            st.rerun()

    st.divider()

    bulk_upload = st.file_uploader(
        "Upload bulk data workbook (.xlsx)",
        type=["xlsx"],
        key="bulk_up_tab"
    )

    data_year = st.selectbox(
        "Data Year (for header parsing)",
        options=years_opt,
        index=years_opt.index(int(view_year)) if int(view_year) in years_opt else years_opt.index(this_year),
        key="bulk_data_year"
    )

    mode = st.radio(
        "Ingest mode",
        options=["Append (add rows)", "Overwrite year + retailer(s) (replace)"],
        index=0,
        horizontal=True,
        key="bulk_mode"
    )

    is_locked = int(data_year) in load_year_locks()
    if is_locked:
        st.error(f"Year {int(data_year)} is locked. Unlock it above to ingest data for this year.")

    c1, c2 = st.columns([1, 3])
    with c1:
        if st.button("Ingest Bulk Workbook", disabled=(bulk_upload is None) or is_locked, key="btn_ingest_bulk"):
            new_rows = read_yow_workbook(bulk_upload, year=int(data_year))

            if mode.startswith("Overwrite"):
                retailers = set(new_rows["Retailer"].dropna().unique().tolist()) if not new_rows.empty else set()
                overwrite_sales_rows(int(data_year), retailers)

            append_sales_to_store(new_rows)
            st.success("Bulk workbook ingested successfully.")
            st.rerun()

    with c2:
        st.caption("Append = adds rows. Overwrite = deletes existing rows for that year + retailer(s) found in the upload, then re-adds.")


def render_seasonality():
        st.subheader("Seasonality (Top 20 seasonal SKUs)")

        if df_all.empty:
            st.info("No sales data yet.")
        else:
            d = df_all.copy()
            d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
            d = d[d["StartDate"].notna()].copy()
            d["Year"] = d["StartDate"].dt.year.astype(int)
            d["Month"] = d["StartDate"].dt.month.astype(int)
            d["MonthP"] = d["StartDate"].dt.to_period("M")

            month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
            month_list = [month_name[i] for i in range(1,13)]

            years = sorted(d["Year"].unique().tolist())

            c1, c2, c3 = st.columns([1, 2, 2])
            with c1:
                basis = st.radio("Basis", options=["Units", "Sales"], index=0, horizontal=True, key="sea_basis")
            with c2:
                mode = st.selectbox("Timeframe", options=["Pick year", "Lookback"], index=0, key="sea_tf_mode")
            with c3:
                # pick year or lookback window
                if mode == "Pick year":
                    pick_year = st.selectbox("Year", options=["All years"] + [str(y) for y in years], index=0, key="sea_year")
                    month_mode = st.radio("Months", options=["All months (Jan–Dec)", "Custom months"], index=0, horizontal=True, key="sea_month_mode")
                    if month_mode == "Custom months":
                        sel_month_names = st.multiselect("Select months", options=month_list, default=month_list, key="sea_months_pick")
                        sel_months = [k for k,v in month_name.items() if v in sel_month_names]
                    else:
                        sel_months = list(range(1,13))
                    # Apply filters
                    d2 = d[d["Month"].isin(sel_months)].copy()
                    if pick_year != "All years":
                        d2 = d2[d2["Year"] == int(pick_year)].copy()
                else:
                    lookback = st.selectbox("Look back", options=["12 months","24 months","36 months","All available"], index=0, key="sea_lookback")
                    if lookback == "All available":
                        d2 = d.copy()
                    else:
                        n = int(lookback.split()[0])
                        months = sorted(d["MonthP"].dropna().unique().tolist())
                        usem = months[-n:] if len(months) >= n else months
                        d2 = d[d["MonthP"].isin(usem)].copy()

            min_units = st.number_input(
                "Minimum total units (within selected timeframe) to include a SKU",
                min_value=0, max_value=1_000_000, value=20, step=5, key="sea_min_units"
            )

            value_col = "Units" if basis == "Units" else "Sales"

            # Monthly totals per SKU (within timeframe)
            m = d2.groupby(["SKU","MonthP"], as_index=False).agg(v=(value_col,"sum"))

            # Seasonality score computed on month-of-year buckets (Jan..Dec) from the same timeframe
            m_y = d2.groupby(["SKU","Month"], as_index=False).agg(v=(value_col,"sum"))
            tot = m_y.groupby("SKU", as_index=False).agg(total=("v","sum"))
            mx = m_y.sort_values("v", ascending=False).groupby("SKU", as_index=False).first().rename(columns={"Month":"PeakMonth","v":"PeakVal"})
            s = tot.merge(mx, on="SKU", how="left")
            s["SeasonalityScore"] = s["PeakVal"] / s["total"].replace(0, np.nan)

            # Filter by units sold in the timeframe (always units)
            units_tot = d2.groupby("SKU", as_index=False).agg(TotalUnits=("Units","sum"))
            s = s.merge(units_tot, on="SKU", how="left").fillna({"TotalUnits": 0})
            s = s[s["TotalUnits"] >= float(min_units)].copy()

            # Vendor labels
            try:
                if isinstance(vmap, pd.DataFrame) and "SKU" in vmap.columns and "Vendor" in vmap.columns:
                    s = s.merge(vmap[["SKU","Vendor"]].drop_duplicates(), on="SKU", how="left")
            except Exception:
                pass

            s = s.sort_values("SeasonalityScore", ascending=False)
            top = s.head(20).copy()
            top["PeakMonthName"] = top["PeakMonth"].map(month_name)

            st.markdown("### Top seasonal SKUs")
            tbl_cols = ["SKU"]
            if "Vendor" in top.columns:
                tbl_cols.append("Vendor")
            tbl_cols += ["PeakMonthName","SeasonalityScore","TotalUnits"]

            tbl = top[tbl_cols].copy().rename(columns={
                "PeakMonthName": "Peak Month",
                "SeasonalityScore": "Seasonality",
                "TotalUnits": "Total Units",
            })
            tbl = tbl.loc[:, ~tbl.columns.duplicated()].copy()

            st.dataframe(
                tbl.style.format({
                    "Seasonality": lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—",
                    "Total Units": fmt_int,
                }),
                use_container_width=True,
                hide_index=True
            )

            st.divider()
            st.markdown("### Seasonal profiles (monthly totals in timeframe)")

            # Create a complete month index for charting, preserving chronological order
            months_all = sorted(d2["MonthP"].dropna().unique().tolist())
            if not months_all:
                st.info("No months found in the selected timeframe.")
                return

            dt_index = pd.PeriodIndex(months_all, freq="M").to_timestamp()

            for _, row in top.iterrows():
                sku0 = row["SKU"]
                vend0 = row.get("Vendor", "")
                peak0 = row.get("PeakMonthName", "")
                score0 = row.get("SeasonalityScore", np.nan)

                title = f"{sku0}"
                if pd.notna(vend0) and str(vend0).strip():
                    title += f" — {vend0}"
                if pd.notna(score0):
                    title += f" | Peak: {peak0} | Seasonality: {score0*100:.1f}%"

                st.markdown(f"**{title}**")

                prof = m[m["SKU"] == sku0][["MonthP","v"]].copy()
                prof["MonthP"] = prof["MonthP"].astype("period[M]")
                prof = prof.set_index("MonthP").reindex(months_all).fillna(0.0)

                chart_df = pd.DataFrame({f"{basis}": prof["v"].to_numpy()}, index=dt_index)
                st.line_chart(chart_df)

def render_runrate():
        st.subheader("Run-Rate Forecast")

        if df.empty:
            st.info("No sales data yet.")
        else:
            window = st.selectbox("Forecast window (weeks)", options=[4, 8, 12], index=0, key="rr_window")
            lookback = st.selectbox("Lookback for avg", options=[4, 8, 12], index=1, key="rr_lookback")
            level = st.selectbox("Level", options=["SKU", "Vendor", "Retailer"], index=0, key="rr_level")

            d = add_week_col(df)
            weeks = last_n_weeks(d, lookback)
            d = d[d["Week"].isin(weeks)].copy()

            if level == "SKU":
                grp = ["Retailer","Vendor","SKU"]
            elif level == "Vendor":
                grp = ["Vendor"]
            else:
                grp = ["Retailer"]

            base = d.groupby(grp + ["Week"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
            units_piv = base.pivot_table(index=grp, columns="Week", values="Units", aggfunc="sum", fill_value=0.0)
            sales_piv = base.pivot_table(index=grp, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0)

            avg_units = nonzero_mean_rowwise(units_piv).fillna(0.0)
            avg_sales = nonzero_mean_rowwise(sales_piv).fillna(0.0)

            out = avg_units.reset_index().rename(columns={0:"AvgWeeklyUnits"})
            out["AvgWeeklySales"] = avg_sales.values
            out["ProjectedUnits"] = out["AvgWeeklyUnits"] * window
            out["ProjectedSales"] = out["AvgWeeklySales"] * window
            out = out.sort_values("ProjectedSales", ascending=False)

            disp = out.copy()
            disp["AvgWeeklyUnits"] = disp["AvgWeeklyUnits"].round(2)
            disp["ProjectedUnits"] = disp["ProjectedUnits"].round(0).astype(int)

            sty = disp.style.format({
                "AvgWeeklyUnits": lambda v: fmt_2(v),
                "AvgWeeklySales": lambda v: fmt_currency(v),
                "ProjectedUnits": lambda v: fmt_int(v),
                "ProjectedSales": lambda v: fmt_currency(v),
            })
            st.dataframe(sty, use_container_width=True, height=_table_height(disp, max_px=1200), hide_index=True)

    # -------------------------
    # Seasonality Heatmap
    # -------------------------



def render_alerts():
        st.subheader("Insights & Alerts")

        if df_all.empty:
            st.info("No sales data yet.")
            return

        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)
        d["MonthP"] = d["StartDate"].dt.to_period("M")

        years = sorted(d["Year"].unique().tolist())
        months = sorted(d["MonthP"].unique().tolist())
        month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
        label_to_period = dict(zip(month_labels, months))

        # --- Period selection ---
        period_mode = st.radio(
            "Period selection",
            options=["Full year (Year vs Year)", "Specific months (A vs B)"],
            index=0,
            horizontal=True,
            key="al_period_mode",
        )

        def _summarize_months(pers: list[pd.Period]) -> str:
            if not pers:
                return "—"
            pers_sorted = sorted(pers)
            labels = [p.to_timestamp().strftime("%b %Y") for p in pers_sorted]
            if len(labels) == 1:
                return labels[0]
            # If they look contiguous, show range; otherwise show count
            try:
                diffs = [(pers_sorted[i+1] - pers_sorted[i]).n for i in range(len(pers_sorted)-1)]
                if diffs and all(int(x) == 1 for x in diffs):
                    return f"{labels[0]}–{labels[-1]}"
            except Exception:
                pass
            return f"{len(labels)} months"

        if period_mode.startswith("Full year"):
            c1, c2 = st.columns(2)
            with c1:
                base_year = st.selectbox("Base Year", options=years, index=0, key="al_base")
            with c2:
                comp_opts = [y for y in years if y != int(base_year)]
                if not comp_opts:
                    st.warning("Only one year of data available. Add another year to compare, or use Specific months.")
                    comp_year = int(base_year)
                else:
                    comp_year = st.selectbox("Comparison Year", options=comp_opts, index=0, key="al_comp")

            a = d[d["Year"] == int(base_year)].copy()
            b = d[d["Year"] == int(comp_year)].copy()
            label_a = str(base_year)
            label_b = str(comp_year)

        else:
            c1, c2 = st.columns(2)
            with c1:
                a_pick = st.multiselect(
                    "Selection A months",
                    options=month_labels,
                    default=month_labels[-1:] if month_labels else [],
                    key="al_a_months",
                )
            with c2:
                b_pick = st.multiselect(
                    "Selection B months",
                    options=month_labels,
                    default=month_labels[-2:-1] if len(month_labels) >= 2 else [],
                    key="al_b_months",
                )

            a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
            b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

            if not a_periods or not b_periods:
                st.info("Pick at least one month in Selection A and Selection B to generate alerts.")
                return

            a = d[d["MonthP"].isin(a_periods)].copy()
            b = d[d["MonthP"].isin(b_periods)].copy()
            label_a = _summarize_months(a_periods)
            label_b = _summarize_months(b_periods)

        basis = st.radio("Basis", options=["Sales", "Units"], index=0, horizontal=True, key="al_basis")
        value_col = "Sales" if basis == "Sales" else "Units"

        insights = []

        # Vendor deltas (worst 5)
        va = a.groupby("Vendor", as_index=False).agg(A=(value_col, "sum"))
        vb = b.groupby("Vendor", as_index=False).agg(B=(value_col, "sum"))
        v = va.merge(vb, on="Vendor", how="outer").fillna(0.0)
        v["Delta"] = v["B"] - v["A"]
        v = v.sort_values("Delta")

        def _fmt(vv):
            return fmt_currency(vv) if value_col == "Sales" else fmt_int(vv)

        for _, row in v.head(5).iterrows():
            if row["Delta"] < 0:
                insights.append(f"🔻 Vendor **{row['Vendor']}** down {_fmt(row['Delta'])} ({label_a} → {label_b}).")

        # Retailer concentration warning (top 1 >= 40%)
        g = b.groupby("Retailer", as_index=False).agg(val=(value_col, "sum")).sort_values("val", ascending=False)
        total = float(g["val"].sum())
        if total > 0 and not g.empty:
            top1_share = float(g.iloc[0]["val"]) / total
            if top1_share >= 0.40:
                insights.append(f"⚠️ Concentration risk: **{g.iloc[0]['Retailer']}** is {top1_share*100:.1f}% of {label_b} ({value_col}).")

        # Growth driven by few SKUs (top10 >= 60% of positive delta)
        sa = a.groupby("SKU", as_index=False).agg(A=(value_col, "sum"))
        sb = b.groupby("SKU", as_index=False).agg(B=(value_col, "sum"))
        sku = sa.merge(sb, on="SKU", how="outer").fillna(0.0)
        sku["Delta"] = sku["B"] - sku["A"]

        pos = sku[sku["Delta"] > 0].sort_values("Delta", ascending=False)
        if not pos.empty:
            top10 = float(pos.head(10)["Delta"].sum())
            total_pos = float(pos["Delta"].sum())
            share = (top10 / total_pos) if total_pos else 0.0
            if share >= 0.60:
                insights.append(f"📈 Growth concentration: top 10 SKUs drive {share*100:.1f}% of positive change ({value_col}) ({label_a} → {label_b}).")

        # Lost SKUs count (had A but not B)
        lost = int(((sku["A"] > 0) & (sku["B"] == 0)).sum())
        if lost:
            insights.append(f"🧯 Lost SKUs: **{lost}** SKUs sold in {label_a} but not in {label_b}.")

        # Year locks notice
        locked = sorted(list(load_year_locks()))
        if locked:
            insights.append(f"🔒 Locked years: {', '.join(str(y) for y in locked)} (bulk ingest blocked).")

        if not insights:
            st.success("No major alerts detected with the current settings.")
        else:
            st.markdown("### Highlights")
            for s in insights:
                st.markdown(f"- {s}")

        with st.expander("Details (tables)", expanded=False):
            st.markdown("**Worst vendors**")
            st.dataframe(
                v.head(15).style.format({"A": _fmt, "B": _fmt, "Delta": _fmt})
                .applymap(lambda v: f"color: {_color(v)};", subset=["Delta"]),
                use_container_width=True,
                hide_index=True
            )

            st.markdown("**Top SKU movers**")
            movers = sku.sort_values("Delta", ascending=False).head(15).copy()
            st.dataframe(
                movers.style.format({"A": _fmt, "B": _fmt, "Delta": _fmt})
                .applymap(lambda v: f"color: {_color(v)};", subset=["Delta"]),
                use_container_width=True,
                hide_index=True
            )


    # Run-Rate Forecast
    # -------------------------


def render_no_sales():
        st.subheader("No Sales SKUs")
        weeks = st.selectbox("Timeframe (weeks)", options=[3,6,8,12], index=0, key="ns_weeks")
        retailers = sorted(vmap["Retailer"].dropna().unique().tolist())
        sel_r = st.selectbox("Retailer", options=["All"] + retailers, index=0, key="ns_retailer")

        if df.empty:
            st.info("No sales data yet.")
        else:
            d2 = df.copy()
            d2["StartDate"] = pd.to_datetime(d2["StartDate"], errors="coerce")
            periods = sorted(d2["StartDate"].dropna().dt.date.unique().tolist())
            use = periods[-weeks:] if len(periods) >= weeks else periods

            if not use:
                st.info("No periods found yet.")
            else:
                sold = d2[d2["StartDate"].dt.date.isin(use)].groupby(["Retailer","SKU"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                ref = vmap[["Retailer","SKU","Vendor","MapOrder"]].copy()
                if sel_r != "All":
                    ref = ref[ref["Retailer"] == sel_r].copy()

                merged = ref.merge(sold, on=["Retailer","SKU"], how="left")
                merged["Units"] = merged["Units"].fillna(0.0)
                merged["Sales"] = merged["Sales"].fillna(0.0)

                nos = merged[(merged["Units"] <= 0) & (merged["Sales"] <= 0)].copy()
                nos["Status"] = f"No sales in last {weeks} weeks"
                nos = nos.sort_values(["Retailer","MapOrder","SKU"], ascending=[True, True, True])

                out = nos[["Retailer","Vendor","SKU","Status"]].copy()
                st.dataframe(out, use_container_width=True, height=_table_height(out, max_px=1400), hide_index=True)


    # -------------------------
    # WoW Exceptions
    # -------------------------

tabs = st.tabs([
    "Totals Dashboard",
    "Top SKUs",
    "Executive Summary",
    "Comparisons",
    "WoW Exceptions",
    
    "SKU Intelligence",
    "Forecasting",
    "Year Summary",
    "Alerts",
    "Data Management",
])
(tab_totals_dash, tab_top_skus, tab_exec, tab_comparisons, tab_wow_exc, tab_sku_intel,
 tab_forecasting, tab_year_summary, tab_data_mgmt_hub, tab_data_mgmt) = tabs













def keep_total_last(df_in: pd.DataFrame, label_col: str) -> pd.DataFrame:
    """After sorting, keep the TOTAL row (if present) pinned to the bottom."""
    try:
        if df_in is None or df_in.empty or label_col not in df_in.columns:
            return df_in
        m = df_in[label_col].astype(str).str.upper().eq("TOTAL")
        if not m.any():
            return df_in
        total = df_in.loc[m].copy()
        rest = df_in.loc[~m].copy()
        return pd.concat([rest, total], ignore_index=True)
    except Exception:
        return df_in


def resolve_week_dates(periods: list, window):
    """
    periods: sorted list of datetime.date representing week start dates.
    window: int weeks or string like "6 months".
    Returns list of week dates to include, ordered ascending.
    """
    if not periods:
        return []
    if isinstance(window, int):
        return periods[-window:] if len(periods) >= window else periods
    if isinstance(window, str) and "month" in window:
        try:
            n = int(window.split()[0])
        except Exception:
            n = 6
        # get last n unique months present in periods
        months = [pd.Timestamp(d).to_period("M") for d in periods]
        uniq = []
        for p in months:
            if p not in uniq:
                uniq.append(p)
        usem = uniq[-n:] if len(uniq) >= n else uniq
        use = [d for d in periods if pd.Timestamp(d).to_period("M") in usem]
        return use
    return periods


def make_totals_tables(base: pd.DataFrame, group_col: str, tf_weeks, avg_weeks):
    if base.empty:
        return pd.DataFrame(), pd.DataFrame()
    base = base.copy()
    base["StartDate"] = pd.to_datetime(base["StartDate"], errors="coerce")
    periods = sorted(base["StartDate"].dropna().dt.date.unique().tolist())
    first_week = periods[0] if periods else None
    if not periods:
        return pd.DataFrame(), pd.DataFrame()

    use = resolve_week_dates(periods, tf_weeks)
    d = base[base["StartDate"].dt.date.isin(use)].copy()
    d["Week"] = d["StartDate"].dt.date

    sales_p = d.pivot_table(index=group_col, columns="Week", values="Sales", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)
    units_p = d.pivot_table(index=group_col, columns="Week", values="Units", aggfunc="sum", fill_value=0.0).reindex(columns=use, fill_value=0.0)

    if len(use) >= 2:
        sales_p["Diff"] = sales_p[use[-1]] - sales_p[use[-2]]
        units_p["Diff"] = units_p[use[-1]] - units_p[use[-2]]
    else:
        sales_p["Diff"] = 0.0
        units_p["Diff"] = 0.0

    # Determine which weeks to average based on selected average window
    current_year = int(pd.to_datetime(base["StartDate"], errors="coerce").dt.year.max() or date.today().year)
    avg_use = resolve_avg_use(avg_weeks, use, current_year)

    # Ignore the very first week of the year (partial week)
    if first_week is not None and avg_use:
        avg_use = [w for w in avg_use if pd.to_datetime(w, errors="coerce").date() != first_week]

    sales_p["Avg"] = sales_p[avg_use].replace(0, np.nan).mean(axis=1) if avg_use else 0.0
    units_p["Avg"] = units_p[avg_use].replace(0, np.nan).mean(axis=1) if avg_use else 0.0

    # Diff vs Avg uses the last week displayed minus Avg
    if use:
        sales_p["Diff vs Avg"] = sales_p[use[-1]] - sales_p["Avg"]
        units_p["Diff vs Avg"] = units_p[use[-1]] - units_p["Avg"]
    else:
        sales_p["Diff vs Avg"] = 0.0
        units_p["Diff vs Avg"] = 0.0

    sales_p = sales_p.sort_index()
    units_p = units_p.sort_index()

    sales_p.loc["TOTAL"] = sales_p.sum(axis=0)
    units_p.loc["TOTAL"] = units_p.sum(axis=0)

    # Recompute TOTAL Avg and Diff vs Avg from totals row values
    if "Avg" in sales_p.columns and use:
        sales_p.loc["TOTAL","Avg"] = sales_p.loc["TOTAL", [c for c in avg_use]].replace(0, np.nan).mean() if avg_use else 0.0
        units_p.loc["TOTAL","Avg"] = units_p.loc["TOTAL", [c for c in avg_use]].replace(0, np.nan).mean() if avg_use else 0.0
        sales_p.loc["TOTAL","Diff vs Avg"] = sales_p.loc["TOTAL", use[-1]] - sales_p.loc["TOTAL","Avg"]
        units_p.loc["TOTAL","Diff vs Avg"] = units_p.loc["TOTAL", use[-1]] - units_p.loc["TOTAL","Avg"]

    def wlab(c):
        try:
            return pd.Timestamp(c).strftime("%m-%d")
        except Exception:
            return c

    sales_p = sales_p.rename(columns={c: wlab(c) for c in sales_p.columns})
    units_p = units_p.rename(columns={c: wlab(c) for c in units_p.columns})

    return sales_p.reset_index(), units_p.reset_index()

# Retailer Totals

with tab_totals_dash:
    st.subheader("Totals Dashboard")

    if df_all.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)
        d["Month"] = d["StartDate"].dt.month.astype(int)

        years = sorted(d["Year"].unique().tolist())
        month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
        month_list = [month_name[i] for i in range(1,13)]

        c1, c2, c3 = st.columns([1, 2, 2])
        with c1:
            year_opt = ["All years"] + [str(y) for y in years]
            pick_year = st.selectbox("Year", options=year_opt, index=0, key="td_year")
        with c2:
            month_mode = st.radio("Months", options=["All months", "Custom months"], index=0, horizontal=True, key="td_month_mode")
            if month_mode == "Custom months":
                sel_month_names = st.multiselect("Select months", options=month_list, default=month_list, key="td_months")
                sel_months = [k for k,v in month_name.items() if v in sel_month_names]
            else:
                sel_months = list(range(1,13))
        with c3:
            group_by = st.selectbox("Group by", options=["Retailer", "Vendor", "SKU"], index=0, key="td_group_by")

        d2 = d[d["Month"].isin(sel_months)].copy()
        if pick_year != "All years":
            d2 = d2[d2["Year"] == int(pick_year)].copy()

        # Optional filters
        f1, f2 = st.columns([2, 2])
        with f1:
            vendor_filter = st.multiselect(
                "Vendor filter (optional)",
                options=sorted([x for x in d2["Vendor"].dropna().unique().tolist() if str(x).strip()]),
                key="td_vendor_filter"
            )
        with f2:
            retailer_filter = st.multiselect(
                "Retailer filter (optional)",
                options=sorted([x for x in d2["Retailer"].dropna().unique().tolist() if str(x).strip()]),
                key="td_retailer_filter"
            )
        if vendor_filter:
            d2 = d2[d2["Vendor"].isin(vendor_filter)]
        if retailer_filter:
            d2 = d2[d2["Retailer"].isin(retailer_filter)]

        if d2.empty:
            st.info("No rows match your filters.")
        else:
            w1, w2, w3 = st.columns([2, 1, 1])
            with w1:
                tf_opt = st.selectbox(
                    "Weeks shown",
                    options=["4 weeks", "6 weeks", "8 weeks", "13 weeks", "26 weeks", "52 weeks", "3 months", "6 months", "12 months", "All available"],
                    index=1,
                    key="td_tf_weeks"
                )
            with w2:
                avg_opt = st.selectbox(
                    "Average window",
                    options=["4 weeks", "6 weeks", "8 weeks", "13 weeks", "26 weeks", "52 weeks"] + month_year_list,

                    index=1,
                    key="td_avg_weeks"
                )
            with w3:
                view_mode = st.selectbox("View", options=["Weekly (with Diff/Avg)", "Summary totals"], index=0, key="td_view_mode")

            def _tf_map(x):
                if x == "All available":
                    return "all"
                if "month" in x:
                    return x
                try:
                    return int(x.split()[0])
                except Exception:
                    return 13

            tf_weeks = _tf_map(tf_opt)

            if view_mode.startswith("Weekly"):
                sales_t, units_t = make_totals_tables(d2, group_by, tf_weeks, avg_opt)
                # Keep alphabetical order for readability
                if not sales_t.empty and group_by in sales_t.columns:
                    sales_t = sales_t.sort_values(group_by, ascending=True, kind="mergesort")
                    sales_t = keep_total_last(sales_t, group_by)
                if not units_t.empty and group_by in units_t.columns:
                    units_t = units_t.sort_values(group_by, ascending=True, kind="mergesort")
                    units_t = keep_total_last(units_t, group_by)

                if sales_t.empty and units_t.empty:
                    st.info("No weekly totals available for the selected filters.")
                else:
                    tabS, tabU = st.tabs(["Sales", "Units"])

                    with tabS:
                        _df = sales_t.copy()

                        def _diff_color(v):
                            try:
                                v = float(v)
                            except Exception:
                                return ""
                            if v > 0:
                                return "color: #2ecc71; font-weight:600;"
                            if v < 0:
                                return "color: #e74c3c; font-weight:600;"
                            return "color: #999999;"

                        diff_cols = [c for c in _df.columns if c in ["Diff", "Diff vs Avg"]]
                        sty = _df.style.format({c: fmt_currency for c in _df.columns if c != group_by})
                        if diff_cols:
                            sty = sty.applymap(lambda v: _diff_color(v), subset=diff_cols)

                        # Bold TOTAL row (if present)
                        try:
                            if group_by in _df.columns:
                                total_mask = _df[group_by].astype(str).str.upper().eq("TOTAL")
                                if total_mask.any():
                                    def _bold_total(row):
                                        return ["font-weight:700;" if str(row.get(group_by,"")).upper()=="TOTAL" else "" for _ in row]
                                    sty = sty.apply(_bold_total, axis=1)
                        except Exception:
                            pass

                        _max_px = 1600 if group_by == "SKU" else 1200
                        st.dataframe(
                            sty,
                            use_container_width=True,
                            hide_index=True,
                            height=_table_height(_df, max_px=_max_px),
                        )
                    with tabU:
                        _df = units_t.copy()

                        def _diff_color(v):
                            try:
                                v = float(v)
                            except Exception:
                                return ""
                            if v > 0:
                                return "color: #2ecc71; font-weight:600;"
                            if v < 0:
                                return "color: #e74c3c; font-weight:600;"
                            return "color: #999999;"

                        diff_cols = [c for c in _df.columns if c in ["Diff", "Diff vs Avg"]]
                        sty = _df.style.format({c: fmt_int for c in _df.columns if c != group_by})
                        if diff_cols:
                            sty = sty.applymap(lambda v: _diff_color(v), subset=diff_cols)

                        # Bold TOTAL row (if present)
                        try:
                            if group_by in _df.columns:
                                total_mask = _df[group_by].astype(str).str.upper().eq("TOTAL")
                                if total_mask.any():
                                    def _bold_total(row):
                                        return ["font-weight:700;" if str(row.get(group_by,"")).upper()=="TOTAL" else "" for _ in row]
                                    sty = sty.apply(_bold_total, axis=1)
                        except Exception:
                            pass

                        _max_px = 1600 if group_by == "SKU" else 1200
                        st.dataframe(
                            sty,
                            use_container_width=True,
                            hide_index=True,
                            height=_table_height(_df, max_px=_max_px),
                        )
            else:
                key = group_by
                agg = d2.groupby(key, as_index=False).agg(
                    Units=("Units","sum"),
                    Sales=("Sales","sum"),
                    SKUs=("SKU","nunique"),
                )

                # Add TOTAL row (always at the bottom) for Retailer/Vendor/SKU views
                try:
                    total = {
                        key: "TOTAL",
                        "Units": float(pd.to_numeric(agg["Units"], errors="coerce").fillna(0).sum()),
                        "Sales": float(pd.to_numeric(agg["Sales"], errors="coerce").fillna(0).sum()),
                        "SKUs": float(d2["SKU"].nunique()),
                    }
                    agg = pd.concat([agg, pd.DataFrame([total])], ignore_index=True)
                except Exception:
                    pass

                # Alphabetical order + force TOTAL last
                agg = agg.sort_values(key, ascending=True, kind="mergesort")
                agg = keep_total_last(agg, key)

                disp = make_unique_columns(agg)
                sty = disp.style.format({"Units": fmt_int, "Sales": fmt_currency, "SKUs": fmt_int})
                # Bold TOTAL row
                try:
                    if key in disp.columns:
                        def _bold_total(row):
                            return ["font-weight:700;" if str(row.get(key,"")).upper()=="TOTAL" else "" for _ in row]
                        sty = sty.apply(_bold_total, axis=1)
                except Exception:
                    pass

                st.dataframe(
                    sty,
                    use_container_width=True,
                    hide_index=True,
                    height=_table_height(disp, max_px=900)
                )




with tab_top_skus:
    st.subheader("Top SKUs (across all retailers)")

    if df_all.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)
        d["Month"] = d["StartDate"].dt.month.astype(int)

        years = sorted(d["Year"].unique().tolist())
        month_name = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
        month_list = [month_name[i] for i in range(1,13)]

        c1, c2, c3, c4, c5 = st.columns([1, 2, 1, 1, 1])
        with c1:
            year_opt = ["All years"] + [str(y) for y in years]
            pick_year = st.selectbox("Year", options=year_opt, index=0, key="ts_year")
        with c2:
            month_mode = st.radio("Months", options=["All months", "Custom months"], index=0, horizontal=True, key="ts_month_mode")
            if month_mode == "Custom months":
                sel_month_names = st.multiselect("Select months", options=month_list, default=month_list, key="ts_months")
                sel_months = [k for k,v in month_name.items() if v in sel_month_names]
            else:
                sel_months = list(range(1,13))
        with c3:
            sort_by = st.selectbox("Rank by", options=["Sales", "Units"], index=0, key="ts_rank_by")
        with c4:
            top_n = st.number_input("Top N", min_value=10, max_value=5000, value=50, step=10, key="ts_topn")

        with c5:
            min_val = st.number_input(
                f"Min {sort_by}",
                min_value=0.0,
                value=0.0,
                step=1.0,
                key="ts_min_val"
            )

        f1, f2 = st.columns([2, 2])
        with f1:
            vendor_filter = st.multiselect(
                "Vendor filter (optional)",
                options=sorted([x for x in d["Vendor"].dropna().unique().tolist() if str(x).strip()]),
                key="ts_vendor_filter"
            )
        with f2:
            retailer_filter = st.multiselect(
                "Retailer filter (optional)",
                options=sorted([x for x in d["Retailer"].dropna().unique().tolist() if str(x).strip()]),
                key="ts_retailer_filter"
            )

        d2 = d[d["Month"].isin(sel_months)].copy()
        if pick_year != "All years":
            d2 = d2[d2["Year"] == int(pick_year)].copy()
        if vendor_filter:
            d2 = d2[d2["Vendor"].isin(vendor_filter)]
        if retailer_filter:
            d2 = d2[d2["Retailer"].isin(retailer_filter)]

        agg = d2.groupby("SKU", as_index=False).agg(
            Units=("Units","sum"),
            Sales=("Sales","sum"),
            Retailers=("Retailer","nunique"),
        )


        # Apply minimum threshold filter (based on Rank by selection)
        if 'min_val' in locals() and min_val and sort_by in agg.columns:
            agg = agg[agg[sort_by].fillna(0) >= float(min_val)].copy()

        if agg.empty:
            st.info("No rows match your filters.")
        else:
            agg = agg.sort_values(sort_by, ascending=False, kind="mergesort").head(int(top_n))
            agg = make_unique_columns(agg)

            st.dataframe(
                agg.style.format({
                    "Units": fmt_int,
                    "Sales": fmt_currency,
                    "Retailers": fmt_int,
                }),
                use_container_width=True,
                hide_index=True,
                height=650
            )

            st.divider()
            st.markdown("### SKU lookup (cross-retailer totals + breakdown)")

            sku_q = st.text_input("Type a SKU to inspect (example: EGLAI1)", value="", key="ts_sku_q").strip()
            if sku_q:
                qn = str(sku_q).strip().upper()
                dd = d2.copy()
                dd["SKU_N"] = dd["SKU"].astype(str).str.strip().str.upper()
                dd = dd[dd["SKU_N"] == qn].copy()

                if dd.empty:
                    st.warning("No matching rows for that SKU in the current filters.")
                else:
                    tot_units = float(dd["Units"].sum())
                    tot_sales = float(dd["Sales"].sum())
                    a, b, c = st.columns([1,1,2])
                    a.metric("Total Units", fmt_int(tot_units))
                    b.metric("Total Sales", fmt_currency(tot_sales))
                    c.caption("Breakdown below is by retailer for the selected year/month filters.")

                    by_ret = dd.groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
                    by_ret = by_ret.sort_values("Sales", ascending=False, kind="mergesort")
                    st.dataframe(
                        by_ret.style.format({"Units": fmt_int, "Sales": fmt_currency}),
                        use_container_width=True,
                        hide_index=True
                    )



with tab_wow_exc:
    st.subheader("WoW Exceptions (Most Recent Week vs Prior Average)")

    if df.empty:
        st.info("No sales data yet.")
    else:
        d0 = add_week_col(df)
        weeks_all = sorted(d0["Week"].dropna().unique().tolist())
        if len(weeks_all) < 2:
            st.info("Not enough weeks loaded yet (need at least 2).")
        else:
            scope = st.selectbox("Scope", options=["All", "Retailer", "Vendor"], index=0, key="wow_scope")

            d1 = d0.copy()
            if scope == "Retailer":
                opts = sorted([x for x in d1["Retailer"].dropna().unique().tolist() if str(x).strip()])
                pick = st.selectbox("Retailer", options=opts, index=0 if opts else 0, key="wow_pick_retailer")
                d1 = d1[d1["Retailer"] == pick].copy()
            elif scope == "Vendor":
                opts = sorted([x for x in d1["Vendor"].dropna().unique().tolist() if str(x).strip()])
                pick = st.selectbox("Vendor", options=opts, index=0 if opts else 0, key="wow_pick_vendor")
                d1 = d1[d1["Vendor"] == pick].copy()

            c1, c2, c3 = st.columns([1.2, 1.2, 2.0])
            with c1:
                # How far back to average (excluding the most recent week)
                n_prior = st.selectbox(
                    "Prior window",
                    options=["4 weeks", "6 weeks", "8 weeks", "12 weeks", "3 months", "6 months", "All prior"],
                    index=1,
                    key="wow_prior_window"
                )
            with c2:
                basis = st.selectbox("Sort basis", options=["Sales", "Units"], index=0, key="wow_sort_basis")
            with c3:
                if scope == "All":
                    display_mode = st.radio(
                        "Display mode",
                        options=["SKU totals (all retailers)", "Break out by retailer"],
                        index=0,
                        horizontal=True,
                        key="wow_display_mode"
                    )
                else:
                    display_mode = "Break out by retailer"

            # Determine most recent week + which prior weeks to use
            d1 = d1[d1["Week"].notna()].copy()
            weeks_all2 = sorted(d1["Week"].dropna().unique().tolist())
            if len(weeks_all2) < 2:
                st.info("Not enough weeks for this selection.")
            else:
                end_week = weeks_all2[-1]
                prior_weeks_all = weeks_all2[:-1]

                def _select_prior(prior_weeks):
                    if n_prior == "All prior":
                        return prior_weeks
                    if "month" in str(n_prior).lower():
                        nmo = int(str(n_prior).split()[0])
                        tmp = d1[d1["Week"].isin(prior_weeks)].copy()
                        tmp["MonthP"] = pd.to_datetime(tmp["StartDate"], errors="coerce").dt.to_period("M")
                        months = sorted(tmp["MonthP"].dropna().unique().tolist())
                        use_months = months[-nmo:] if len(months) >= nmo else months
                        wk = sorted(tmp[tmp["MonthP"].isin(use_months)]["Week"].dropna().unique().tolist())
                        return wk
                    try:
                        n = int(str(n_prior).split()[0])
                    except Exception:
                        n = 6
                    return prior_weeks[-n:] if len(prior_weeks) >= n else prior_weeks

                prior_weeks = _select_prior(prior_weeks_all)
                if not prior_weeks:
                    st.info("No prior weeks in the selected window.")
                else:
                    if display_mode.startswith("SKU totals"):
                        group_cols = ["SKU"]
                        # helpful extra columns
                        extra_aggs = {"Vendor": ("Vendor", lambda s: s.dropna().astype(str).str.strip().iloc[0] if len(s.dropna()) else ""),
                                      "Retailers": ("Retailer", "nunique")}
                    else:
                        group_cols = ["Retailer", "Vendor", "SKU"]
                        extra_aggs = {}

                    dd = d1.copy()

                    # Aggregate to weekly grain for each group
                    g = dd.groupby(group_cols + ["Week"], as_index=False).agg(
                        Units=("Units", "sum"),
                        Sales=("Sales", "sum"),
                    )

                    # Split into end week and prior weeks
                    end = g[g["Week"] == end_week].copy()
                    base = g[g["Week"].isin(prior_weeks)].copy()

                    base_avg = base.groupby(group_cols, as_index=False).agg(
                        Units_Base=("Units", "mean"),
                        Sales_Base=("Sales", "mean"),
                    )
                    end_sum = end.groupby(group_cols, as_index=False).agg(
                        Units_End=("Units", "sum"),
                        Sales_End=("Sales", "sum"),
                    )

                    t = end_sum.merge(base_avg, on=group_cols, how="outer").fillna(0.0)
                    t["Units_Diff"] = t["Units_End"] - t["Units_Base"]
                    t["Sales_Diff"] = t["Sales_End"] - t["Sales_Base"]
                    t["Units_% Diff"] = t["Units_Diff"] / t["Units_Base"].replace(0, np.nan)
                    t["Sales_% Diff"] = t["Sales_Diff"] / t["Sales_Base"].replace(0, np.nan)

                    # Add vendor / retailer coverage when in SKU totals mode
                    if display_mode.startswith("SKU totals"):
                        cov = dd.groupby("SKU", as_index=False).agg(
                            Vendor=("Vendor", lambda s: s.dropna().astype(str).str.strip().iloc[0] if len(s.dropna()) else ""),
                            Retailers=("Retailer", "nunique")
                        )
                        t = t.merge(cov, on="SKU", how="left")

                    sort_col = "Sales_Diff" if basis == "Sales" else "Units_Diff"
                    t = t.sort_values(sort_col, ascending=True, kind="mergesort")  # show biggest negatives first

                    # Keep useful column order
                    if display_mode.startswith("SKU totals"):
                        cols = ["SKU", "Vendor", "Retailers",
                                "Units_Base", "Units_End", "Units_Diff", "Units_% Diff",
                                "Sales_Base", "Sales_End", "Sales_Diff", "Sales_% Diff"]
                    else:
                        cols = ["Retailer", "Vendor", "SKU",
                                "Units_Base", "Units_End", "Units_Diff", "Units_% Diff",
                                "Sales_Base", "Sales_End", "Sales_Diff", "Sales_% Diff"]
                    cols = [c for c in cols if c in t.columns]
                    t = t[cols].copy()

                    # Totals row at bottom for quick reference
                    try:
                        total = {c: "" for c in t.columns}
                        first = t.columns[0]
                        total[first] = "TOTAL"
                        for c in t.columns:
                            if c in {"SKU","Vendor","Retailer"}:
                                continue
                            if c == "Retailers":
                                total[c] = float(dd["Retailer"].nunique())
                            else:
                                total[c] = float(pd.to_numeric(t[c], errors="coerce").fillna(0).sum())
                        t = pd.concat([t, pd.DataFrame([total])], ignore_index=True)
                    except Exception:
                        pass

                    # Styling
                    disp = make_unique_columns(t)

                    def _diff_color(v):
                        try:
                            v = float(v)
                        except Exception:
                            return ""
                        if v > 0:
                            return "color: #2ecc71; font-weight:600;"
                        if v < 0:
                            return "color: #e74c3c; font-weight:600;"
                        return "color: #999999;"

                    sty = disp.style.format({
                        "Units_Base": fmt_int,
                        "Units_End": fmt_int,
                        "Units_Diff": fmt_int_signed,
                        "Units_% Diff": lambda v: f"{(v*100):.1f}%" if pd.notna(v) else "—",
                        "Sales_Base": fmt_currency,
                        "Sales_End": fmt_currency,
                        "Sales_Diff": fmt_currency_signed,
                        "Sales_% Diff": lambda v: f"{(v*100):.1f}%" if pd.notna(v) else "—",
                        "Retailers": fmt_int,
                    })

                    for c in ["Units_Diff", "Sales_Diff"]:
                        if c in disp.columns:
                            sty = sty.applymap(lambda v: _diff_color(v), subset=[c])

                    # Bold TOTAL row (if present)
                    try:
                        first = disp.columns[0]
                        if first in disp.columns:
                            def _bold_total(row):
                                return ["font-weight:700;" if str(row.get(first,"")).upper()=="TOTAL" else "" for _ in row]
                            sty = sty.apply(_bold_total, axis=1)
                    except Exception:
                        pass

                    st.caption(
                        f"Comparing most recent week ({pd.Timestamp(end_week).strftime('%m-%d')}) "
                        f"to the average of prior {len(prior_weeks)} week(s): "
                        + ", ".join([pd.Timestamp(w).strftime('%m-%d') for w in prior_weeks])
                    )
                    st.dataframe(sty, use_container_width=True, height=_table_height(disp, max_px=1200), hide_index=True)



with tab_comparisons:
    st.subheader("Comparisons")

    view = st.selectbox("View", options=["Retailer / Vendor Comparison", "SKU Comparison"], index=0, key="cmp_view")
    if view == "Retailer / Vendor Comparison":
        render_comparison_retailer_vendor()
        a_key, b_key = "cmp_a_months", "cmp_b_months"
    else:
        render_comparison_sku()
        a_key, b_key = "skucmp_a_months", "skucmp_b_months"

    # -------------------------
    # Top SKU movers (combined across all retailers)
    # -------------------------
    st.divider()
    st.markdown("### Top SKU movers (all retailers combined)")

    try:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["MonthP"] = d["StartDate"].dt.to_period("M")

        months = sorted(d["MonthP"].unique().tolist())
        month_labels = [m.to_timestamp().strftime("%B %Y") for m in months]
        label_to_period = dict(zip(month_labels, months))

        a_pick = st.session_state.get(a_key, [])
        b_pick = st.session_state.get(b_key, [])

        a_periods = [label_to_period[x] for x in a_pick if x in label_to_period]
        b_periods = [label_to_period[x] for x in b_pick if x in label_to_period]

        if (not a_periods) or (not b_periods):
            st.caption("Pick Selection A and Selection B above to populate movers.")
        else:
            da = d[d["MonthP"].isin(a_periods)]
            db = d[d["MonthP"].isin(b_periods)]

            ga = da.groupby("SKU", as_index=False).agg(Units_A=("Units","sum"), Sales_A=("Sales","sum"))
            gb = db.groupby("SKU", as_index=False).agg(Units_B=("Units","sum"), Sales_B=("Sales","sum"))

            out = ga.merge(gb, on="SKU", how="outer").fillna(0.0)
            out["Units_Diff"] = out["Units_A"] - out["Units_B"]
            out["Sales_Diff"] = out["Sales_A"] - out["Sales_B"]

            # Add vendor + retailer coverage for context
            cov = d.groupby("SKU", as_index=False).agg(
                Vendor=("Vendor", lambda s: (s.dropna().astype(str).str.strip().replace("", np.nan).dropna().iloc[0] if s.dropna().astype(str).str.strip().replace("", np.nan).dropna().shape[0] else "Unmapped")),
                Retailers=("Retailer","nunique"),
            )
            out = out.merge(cov, on="SKU", how="left")

            inc = out.sort_values("Sales_Diff", ascending=False, kind="mergesort").head(10)
            dec = out.sort_values("Sales_Diff", ascending=True, kind="mergesort").head(10)

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("#### Top 10 increased (by Sales Δ)")
                disp = inc[["SKU","Vendor","Retailers","Units_A","Units_B","Units_Diff","Sales_A","Sales_B","Sales_Diff"]].copy()
                st.dataframe(
                    disp.style.format({"Retailers": fmt_int, "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int_signed,
                                       "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency_signed}),
                    use_container_width=True,
                    hide_index=True,
                    height=_table_height(disp, max_px=520),
                )
            with c2:
                st.markdown("#### Top 10 decreased (by Sales Δ)")
                disp = dec[["SKU","Vendor","Retailers","Units_A","Units_B","Units_Diff","Sales_A","Sales_B","Sales_Diff"]].copy()
                st.dataframe(
                    disp.style.format({"Retailers": fmt_int, "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int_signed,
                                       "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency_signed}),
                    use_container_width=True,
                    hide_index=True,
                    height=_table_height(disp, max_px=520),
                )
    except Exception:
        st.caption("Movers will appear after you pick Selection A and Selection B above.")


with tab_sku_intel:
    st.subheader("SKU Intelligence")
    view = st.selectbox("View", options=["SKU Health Score", "Lost Sales Detector"], index=0, key="sku_intel_view")
    if view == "SKU Health Score":
        render_sku_health()
    else:
        render_lost_sales()

with tab_forecasting:
    st.subheader("Forecasting")
    view = st.selectbox("View", options=["Run-Rate Forecast", "Seasonality"], index=0, key="fc_view")
    if view == "Run-Rate Forecast":
        render_runrate()
    else:
        render_seasonality()

with tab_data_mgmt_hub:
    st.subheader("Alerts")
    view = st.selectbox("View", options=["Insights & Alerts", "No Sales SKUs"], index=0, key="alerts_view")
    if view == "Insights & Alerts":
        render_alerts()
    else:
        render_no_sales()

with tab_data_mgmt:
    st.subheader("Data Management")
    view = st.selectbox("Section", options=["Data Inventory", "Edit Vendor Map", "Backup / Restore", "Bulk Data Upload"], index=0, key="dm_view")
    if view == "Data Inventory":
        render_data_inventory()
    elif view == "Edit Vendor Map":
        render_edit_vendor_map()
    elif view == "Backup / Restore":
        render_backup_restore()
    else:
        render_bulk_data_upload()

with tab_year_summary:
    st.subheader("Year Summary")

    if df_all.empty:
        st.info("No sales data yet.")
    else:
        d = df_all.copy()
        d["StartDate"] = pd.to_datetime(d["StartDate"], errors="coerce")
        d = d[d["StartDate"].notna()].copy()
        d["Year"] = d["StartDate"].dt.year.astype(int)

        years = sorted(d["Year"].unique().tolist())
        if not years:
            st.info("No dated rows available.")
            st.stop()

        current_year = int(max(years))
        prior_year = int(years[-2]) if len(years) >= 2 else None

        basis = st.radio("Basis", options=["Sales", "Units"], index=0, horizontal=True, key="ys_basis_clean")
        value_col = "Sales" if basis == "Sales" else "Units"

        cur = d[d["Year"] == current_year].copy()
        prv = d[d["Year"] == prior_year].copy() if prior_year is not None else d.iloc[0:0].copy()

        def _sum(df_, col):
            return float(df_[col].sum()) if df_ is not None and not df_.empty else 0.0

        # -------------------------
        # KPIs (always latest two years)
        # -------------------------
        st.markdown("### KPIs (latest two years)")

        uC, sC = _sum(cur, "Units"), _sum(cur, "Sales")
        uP, sP = (_sum(prv, "Units") if prior_year is not None else 0.0), (_sum(prv, "Sales") if prior_year is not None else 0.0)

        uD = uC - uP
        sD = sC - sP
        uPct = (uD / uP) if uP else np.nan
        sPct = (sD / sP) if sP else np.nan

        # All-years context (vs average of all years)
        yr_tot = d.groupby("Year", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Year")
        avg_units = float(yr_tot["Units"].mean()) if not yr_tot.empty else 0.0
        avg_sales = float(yr_tot["Sales"].mean()) if not yr_tot.empty else 0.0
        delta_vs_avg_units = uC - avg_units
        delta_vs_avg_sales = sC - avg_sales

        k1, k2, k3, k4, k5, k6 = st.columns(6)
        k1.metric(f"Units ({current_year})", fmt_int(uC))
        if prior_year is not None:
            k2.metric(f"Units ({prior_year})", fmt_int(uP), delta=f"{fmt_int(uD)} ({uPct*100:.1f}%)" if pd.notna(uPct) else fmt_int(uD))
        else:
            k2.metric("Units (prior)", "—")

        k3.metric(f"Sales ({current_year})", fmt_currency(sC))
        if prior_year is not None:
            k4.metric(f"Sales ({prior_year})", fmt_currency(sP), delta=f"{fmt_currency_signed(sD)} ({sPct*100:.1f}%)" if pd.notna(sPct) else fmt_currency_signed(sD))
        else:
            k4.metric("Sales (prior)", "—")

        k5.markdown(
            f"**Δ Units vs all-years avg**<br><span style='color:{_color(delta_vs_avg_units)}; font-weight:700;'>{fmt_int(delta_vs_avg_units)}</span>",
            unsafe_allow_html=True
        )
        k6.markdown(
            f"**Δ Sales vs all-years avg**<br><span style='color:{_color(delta_vs_avg_sales)}; font-weight:700;'>{fmt_currency_signed(delta_vs_avg_sales)}</span>",
            unsafe_allow_html=True
        )

        # -------------------------
        # YoY driver breakdown (latest two years)
        # -------------------------
        if prior_year is not None:
            st.markdown("### YoY driver breakdown (latest two years)")

            a = prv
            b = cur

            sku_a = a.groupby("SKU", as_index=False).agg(Units_A=("Units", "sum"), Sales_A=("Sales", "sum"))
            sku_b = b.groupby("SKU", as_index=False).agg(Units_B=("Units", "sum"), Sales_B=("Sales", "sum"))
            sku = sku_a.merge(sku_b, on="SKU", how="outer").fillna(0.0)
            sku["Δ Units"] = sku["Units_B"] - sku["Units_A"]
            sku["Δ Sales"] = sku["Sales_B"] - sku["Sales_A"]

            mover_col = "Δ Sales" if value_col == "Sales" else "Δ Units"
            tn = sku.sort_values(mover_col, ascending=False).head(10).copy()
            tn2 = sku.sort_values(mover_col, ascending=True).head(10).copy()

            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Top 10 Increases**")
                tn_disp = tn[["SKU", "Units_A", "Units_B", "Δ Units", "Sales_A", "Sales_B", "Δ Sales"]]
                st.dataframe(
                    tn_disp.style.format({
                        "Units_A": fmt_int, "Units_B": fmt_int, "Δ Units": fmt_int,
                        "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Δ Sales": fmt_currency,
                    }).applymap(lambda v: f"color: {_color(v)};", subset=["Δ Units", "Δ Sales"]),
                    use_container_width=True,
                    height=_table_height(tn_disp, max_px=700),
                    hide_index=True
                )
            with c2:
                st.markdown("**Top 10 Decreases**")
                tn2_disp = tn2[["SKU", "Units_A", "Units_B", "Δ Units", "Sales_A", "Sales_B", "Δ Sales"]]
                st.dataframe(
                    tn2_disp.style.format({
                        "Units_A": fmt_int, "Units_B": fmt_int, "Δ Units": fmt_int,
                        "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Δ Sales": fmt_currency,
                    }).applymap(lambda v: f"color: {_color(v)};", subset=["Δ Units", "Δ Sales"]),
                    use_container_width=True,
                    height=_table_height(tn2_disp, max_px=700),
                    hide_index=True
                )

        # -------------------------
        # Concentration risk (ALL YEARS, Retailer + Vendor)
        # -------------------------
        st.markdown("### Concentration risk (all years)")

        def _top_share(df_year, group_col, topn):
            g = df_year.groupby(group_col, as_index=False).agg(val=(value_col, "sum"))
            total = float(g["val"].sum())
            if total <= 0:
                return 0.0
            return float(g.sort_values("val", ascending=False).head(topn)["val"].sum()) / total

        conc_rows = []
        for y in years:
            dy = d[d["Year"] == int(y)].copy()
            conc_rows.append({
                "Year": int(y),
                "Top 1 Retailer %": _top_share(dy, "Retailer", 1),
                "Top 3 Retailers %": _top_share(dy, "Retailer", 3),
                "Top 5 Retailers %": _top_share(dy, "Retailer", 5),
                "Top 1 Vendor %": _top_share(dy, "Vendor", 1),
                "Top 3 Vendors %": _top_share(dy, "Vendor", 3),
                "Top 5 Vendors %": _top_share(dy, "Vendor", 5),
            })

        conc = pd.DataFrame(conc_rows).sort_values("Year")
        st.dataframe(
            conc.style.format({c: (lambda v: f"{v*100:.1f}%") for c in conc.columns if c != "Year"}),
            use_container_width=True,
            hide_index=True
        )

        # -------------------------
        # Accordion breakdown by year (click to expand/collapse)
        # -------------------------
        st.caption("Click a year to expand the Top 1 / Top 3 / Top 5 breakdown. Click again to collapse.")

        def _top_list(df_year, group_col, topn):
            g = df_year.groupby(group_col, as_index=False).agg(val=(value_col, "sum"))
            total = float(g["val"].sum())
            if total <= 0:
                return pd.DataFrame(columns=[group_col, value_col, "Share"])
            g = g.sort_values("val", ascending=False).head(topn).copy()
            g["Share"] = g["val"] / total
            g = g.rename(columns={"val": value_col})
            return g[[group_col, value_col, "Share"]]

        for y in years:
            dy = d[d["Year"] == int(y)].copy()
            with st.expander(f"{int(y)} — Top 1 / Top 3 / Top 5 breakdown", expanded=False):
                c1, c2 = st.columns(2)

                with c1:
                    st.markdown("**Retailers**")
                    t1 = _top_list(dy, "Retailer", 1)
                    t3 = _top_list(dy, "Retailer", 3)
                    t5 = _top_list(dy, "Retailer", 5)

                    st.markdown("Top 1 Retailer")
                    st.dataframe(
                        t1.style.format({value_col: (fmt_currency if value_col == "Sales" else fmt_int),
                                         "Share": (lambda v: f"{v*100:.1f}%")}),
                        use_container_width=True,
                        hide_index=True
                    )
                    st.markdown("Top 3 Retailers")
                    st.dataframe(
                        t3.style.format({value_col: (fmt_currency if value_col == "Sales" else fmt_int),
                                         "Share": (lambda v: f"{v*100:.1f}%")}),
                        use_container_width=True,
                        hide_index=True
                    )
                    st.markdown("Top 5 Retailers")
                    st.dataframe(
                        t5.style.format({value_col: (fmt_currency if value_col == "Sales" else fmt_int),
                                         "Share": (lambda v: f"{v*100:.1f}%")}),
                        use_container_width=True,
                        hide_index=True
                    )

                with c2:
                    st.markdown("**Vendors**")
                    v1 = _top_list(dy, "Vendor", 1)
                    v3 = _top_list(dy, "Vendor", 3)
                    v5 = _top_list(dy, "Vendor", 5)

                    st.markdown("Top 1 Vendor")
                    st.dataframe(
                        v1.style.format({value_col: (fmt_currency if value_col == "Sales" else fmt_int),
                                         "Share": (lambda v: f"{v*100:.1f}%")}),
                        use_container_width=True,
                        hide_index=True
                    )
                    st.markdown("Top 3 Vendors")
                    st.dataframe(
                        v3.style.format({value_col: (fmt_currency if value_col == "Sales" else fmt_int),
                                         "Share": (lambda v: f"{v*100:.1f}%")}),
                        use_container_width=True,
                        hide_index=True
                    )
                    st.markdown("Top 5 Vendors")
                    st.dataframe(
                        v5.style.format({value_col: (fmt_currency if value_col == "Sales" else fmt_int),
                                         "Share": (lambda v: f"{v*100:.1f}%")}),
                        use_container_width=True,
                        hide_index=True
                    )

        # -------------------------
        # Retailer summary (year pick ONLY here)
        # -------------------------
        st.markdown("### Retailer summary (pick years here)")
        r1, r2 = st.columns(2)
        with r1:
            r_base = st.selectbox("Retailer Base Year", options=years, index=max(0, len(years) - 2), key="ys_r_base_clean")
        with r2:
            r_comp = st.selectbox("Retailer Comparison Year", options=years, index=len(years) - 1, key="ys_r_comp_clean")

        ra = d[d["Year"] == int(r_base)].groupby("Retailer", as_index=False).agg(Units_A=("Units", "sum"), Sales_A=("Sales", "sum"))
        rb = d[d["Year"] == int(r_comp)].groupby("Retailer", as_index=False).agg(Units_B=("Units", "sum"), Sales_B=("Sales", "sum"))
        r = ra.merge(rb, on="Retailer", how="outer").fillna(0.0)
        r["Units_Diff"] = r["Units_B"] - r["Units_A"]
        r["Sales_Diff"] = r["Sales_B"] - r["Sales_A"]
        r["Units_%"] = r["Units_Diff"] / r["Units_A"].replace(0, np.nan)
        r["Sales_%"] = r["Sales_Diff"] / r["Sales_A"].replace(0, np.nan)

        rsort = st.selectbox("Sort retailer table by", ["Sales_Diff", "Units_Diff", "Sales_B", "Sales_A"], key="ys_r_sort_clean")
        r = r.sort_values(rsort, ascending=False, kind="mergesort")
        r_disp = r[["Retailer", "Units_A", "Units_B", "Units_Diff", "Units_%", "Sales_A", "Sales_B", "Sales_Diff", "Sales_%"]]
        r_sty = r_disp.style.format({
            "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int,
            "Units_%": (lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"),
            "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency,
            "Sales_%": (lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"),
        }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff", "Sales_Diff"])
        st.dataframe(r_sty, use_container_width=True, height=_table_height(r_disp, max_px=1100), hide_index=True)

        # -------------------------
        # Vendor summary (year pick ONLY here)
        # -------------------------
        st.markdown("### Vendor summary (pick years here)")
        v1, v2 = st.columns(2)
        with v1:
            v_base = st.selectbox("Vendor Base Year", options=years, index=max(0, len(years) - 2), key="ys_v_base_clean")
        with v2:
            v_comp = st.selectbox("Vendor Comparison Year", options=years, index=len(years) - 1, key="ys_v_comp_clean")

        va = d[d["Year"] == int(v_base)].groupby("Vendor", as_index=False).agg(Units_A=("Units", "sum"), Sales_A=("Sales", "sum"))
        vb = d[d["Year"] == int(v_comp)].groupby("Vendor", as_index=False).agg(Units_B=("Units", "sum"), Sales_B=("Sales", "sum"))
        v = va.merge(vb, on="Vendor", how="outer").fillna(0.0)
        v["Units_Diff"] = v["Units_B"] - v["Units_A"]
        v["Sales_Diff"] = v["Sales_B"] - v["Sales_A"]
        v["Units_%"] = v["Units_Diff"] / v["Units_A"].replace(0, np.nan)
        v["Sales_%"] = v["Sales_Diff"] / v["Sales_A"].replace(0, np.nan)

        vsort = st.selectbox("Sort vendor table by", ["Sales_Diff", "Units_Diff", "Sales_B", "Sales_A"], key="ys_v_sort_clean")
        v = v.sort_values(vsort, ascending=False, kind="mergesort")
        v_disp = v[["Vendor", "Units_A", "Units_B", "Units_Diff", "Units_%", "Sales_A", "Sales_B", "Sales_Diff", "Sales_%"]]
        v_sty = v_disp.style.format({
            "Units_A": fmt_int, "Units_B": fmt_int, "Units_Diff": fmt_int,
            "Units_%": (lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"),
            "Sales_A": fmt_currency, "Sales_B": fmt_currency, "Sales_Diff": fmt_currency,
            "Sales_%": (lambda v: f"{v*100:.1f}%" if pd.notna(v) else "—"),
        }).applymap(lambda v: f"color: {_color(v)};", subset=["Units_Diff", "Sales_Diff"])
        st.dataframe(v_sty, use_container_width=True, height=_table_height(v_disp, max_px=1100), hide_index=True)


# -------------------------
# Data Inventory
# -------------------------


with tab_exec:
    st.subheader("Executive Summary")

    scope = st.selectbox("Scope", options=["All", "Retailer", "Vendor"], index=0, key="ex_scope")

    # Filter scope
    if scope == "Retailer":
        opts = sorted([x for x in df_all["Retailer"].dropna().unique().tolist() if str(x).strip()])
        pick = st.selectbox("Retailer", options=opts, index=0 if opts else 0, key="ex_pick_r")
        d_base = df_all[df_all["Retailer"] == pick].copy()
        title = f"Executive Summary - {pick}"
    elif scope == "Vendor":
        opts = sorted([x for x in df_all["Vendor"].dropna().unique().tolist() if str(x).strip()])
        pick = st.selectbox("Vendor", options=opts, index=0 if opts else 0, key="ex_pick_v")
        d_base = df_all[df_all["Vendor"] == pick].copy()
        title = f"Executive Summary - {pick}"
    else:
        d_base = df_all.copy()
        title = "Executive Summary - All Retailers"

    d_base["StartDate"] = pd.to_datetime(d_base["StartDate"], errors="coerce")
    d_base = d_base[d_base["StartDate"].notna()].copy()
    if d_base.empty:
        st.info("No rows in the selected scope/year.")
        st.stop()

    d_base["Year"] = d_base["StartDate"].dt.year.astype(int)
    years = sorted(d_base["Year"].unique().tolist())
    pick_year = st.selectbox("Year", options=years, index=(len(years)-1 if years else 0), key="ex_year_pick")
    d = d_base[d_base["Year"] == int(pick_year)].copy()

    st.caption(title)

    # KPI row
    m = wow_mom_metrics(d)
    cols = st.columns(6)
    cols[0].metric("Units", fmt_int(m["total_units"]))
    cols[1].metric("Sales", fmt_currency(m["total_sales"]))
    cols[2].markdown(f"<div style='color:{_color(m['wow_units'])}; font-weight:600;'>WoW Units: {fmt_int(m['wow_units']) if m['wow_units'] is not None else '—'}</div>", unsafe_allow_html=True)
    cols[3].markdown(f"<div style='color:{_color(m['wow_sales'])}; font-weight:600;'>WoW Sales: {fmt_currency(m['wow_sales']) if m['wow_sales'] is not None else '—'}</div>", unsafe_allow_html=True)
    cols[4].markdown(f"<div style='color:{_color(m['mom_units'])}; font-weight:600;'>MoM Units: {fmt_int(m['mom_units']) if m['mom_units'] is not None else '—'}</div>", unsafe_allow_html=True)
    cols[5].markdown(f"<div style='color:{_color(m['mom_sales'])}; font-weight:600;'>MoM Sales: {fmt_currency(m['mom_sales']) if m['mom_sales'] is not None else '—'}</div>", unsafe_allow_html=True)

    st.divider()

    # When Scope = ALL: show one line per SKU (combined across all retailers)
    if scope == "All":
        sku = d.groupby("SKU", as_index=False).agg(
            Vendor=("Vendor", lambda s: (s.dropna().astype(str).str.strip().replace("", np.nan).dropna().iloc[0] if s.dropna().astype(str).str.strip().replace("", np.nan).dropna().shape[0] else "Unmapped")),
            Retailers=("Retailer", "nunique"),
            TotalUnits=("Units", "sum"),
            TotalSales=("Sales", "sum"),
        )
        sku = sku.sort_values("TotalSales", ascending=False, kind="mergesort")

        disp = sku[["SKU","Vendor","Retailers","TotalUnits","TotalSales"]].copy()
        st.markdown("### SKU totals (all retailers combined)")
        st.dataframe(
            disp.style.format({"Retailers": fmt_int, "TotalUnits": fmt_int, "TotalSales": fmt_currency}),
            use_container_width=True,
            hide_index=True,
            height=_table_height(disp, max_px=1100),
        )

    else:
        # Monthly totals table (keep as-is)
        d2 = d.copy()
        d2["MonthP"] = d2["StartDate"].dt.to_period("M")
        mon = d2.groupby("MonthP", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("MonthP")
        if not mon.empty:
            mon["Month"] = mon["MonthP"].map(month_label)
            mon = mon[["Month","Units","Sales"]]
            st.markdown("### Monthly totals")
            st.dataframe(
                mon.style.format({"Units": fmt_int, "Sales": fmt_currency}),
                use_container_width=True,
                height=_table_height(mon, max_px=800),
                hide_index=True
            )

        # Mix table (keep as-is)
        if scope == "Retailer":
            mix = d.groupby("Vendor", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
            mix = mix[(mix["Units"].fillna(0) > 0) | (mix["Sales"].fillna(0) > 0)]
            total_u = float(mix["Units"].sum()) if not mix.empty else 0.0
            total_s = float(mix["Sales"].sum()) if not mix.empty else 0.0
            mix["% Units"] = mix["Units"].apply(lambda v: (v/total_u) if total_u else 0.0)
            mix["% Sales"] = mix["Sales"].apply(lambda v: (v/total_s) if total_s else 0.0)
            mix = mix.sort_values("% Sales", ascending=False, kind="mergesort")
            st.markdown("### Vendor mix")
        else:
            mix = d.groupby("Retailer", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
            mix = mix[(mix["Units"].fillna(0) > 0) | (mix["Sales"].fillna(0) > 0)]
            total_u = float(mix["Units"].sum()) if not mix.empty else 0.0
            total_s = float(mix["Sales"].sum()) if not mix.empty else 0.0
            mix["% Units"] = mix["Units"].apply(lambda v: (v/total_u) if total_u else 0.0)
            mix["% Sales"] = mix["Sales"].apply(lambda v: (v/total_s) if total_s else 0.0)
            mix = mix.sort_values("% Sales", ascending=False, kind="mergesort")
            st.markdown("### Retailer mix")

        st.dataframe(
            mix.style.format({"Units": fmt_int, "Sales": fmt_currency, "% Units": lambda v: f"{v*100:.1f}%", "% Sales": lambda v: f"{v*100:.1f}%"}),
            use_container_width=True,
            height=_table_height(mix, max_px=900),
            hide_index=True
        )

        st.divider()

        # Top / Bottom SKUs (keep the same idea as before)
        sold = d.groupby(["SKU","Vendor"], as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum"))
        sold = sold[(sold["Units"].fillna(0) > 0) | (sold["Sales"].fillna(0) > 0)].copy()

        left, right = st.columns(2)
        with left:
            st.markdown("### Top 10 SKUs (by Sales)")
            top10 = sold.sort_values("Sales", ascending=False, kind="mergesort").head(10)[["SKU","Vendor","Units","Sales"]]
            st.dataframe(top10.style.format({"Units": fmt_int, "Sales": fmt_currency}), use_container_width=True, hide_index=True, height=_table_height(top10, max_px=520))
        with right:
            st.markdown("### Bottom 10 SKUs (by Sales)")
            bot10 = sold.sort_values("Sales", ascending=True, kind="mergesort").head(10)[["SKU","Vendor","Units","Sales"]]
            st.dataframe(bot10.style.format({"Units": fmt_int, "Sales": fmt_currency}), use_container_width=True, hide_index=True, height=_table_height(bot10, max_px=520))

    st.divider()
    st.subheader("Yearly totals (all years)")

    yt = d_base.groupby("Year", as_index=False).agg(Units=("Units","sum"), Sales=("Sales","sum")).sort_values("Year")
    st.dataframe(
        yt.style.format({"Units": fmt_int, "Sales": fmt_currency}),
        use_container_width=True,
        hide_index=True
    )